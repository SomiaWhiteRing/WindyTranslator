from __future__ import annotations

import csv
import json
import os
import re
import shutil
import subprocess
import sys
import unicodedata
import uuid
from collections import Counter, OrderedDict, defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path, PureWindowsPath
from typing import Callable, Iterable, Iterator, Mapping, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from qa_core_legacy import (
    APP_NAME as LEGACY_APP_NAME,
    BACKUP_DIRNAME,
    CONTROL_CODE_RE,
    DataRecord,
    DictionaryEntry,
    DictionaryWarning,
    NON_SPEAKER_IDS,
    ParsedOccurrence,
    QAError,
    SourceLoadResult,
    _column_letter,
    _find_header,
    _normalize_header,
    _no_duplicate_object,
    _safe_windows_file_key,
    _worksheet_dimensions,
    fullwidth_units,
    is_effective_speaker,
    load_dictionary,
    load_dictionaries,
    load_excel_records as legacy_load_excel_records,
    load_json_records,
    longest_non_overlapping_matches,
    normalize_newlines,
    open_file,
    strip_control_codes,
)

APP_NAME = "RPG制作大师校对工具"
TRANSLATION_SHEET = "Translation"
NON_FACE_IDS = {"", "NARRATION", "SYSTEM", "_ERASE_FACE_"}
MAP_RE = re.compile(r"(?:^|[\\/])Map\d+\.txt$", re.I)
COMMON_EVENT_RE = re.compile(r"common\s*_?events?|commons?", re.I)
RE_MARKER_LINE = re.compile(r"^\s*#([^#]+)#")
RE_FACE = re.compile(r"^\s*\{{2,}.*?Select Face Graphic:(.*?)\}{2,}\s*$", re.I)
RE_PAGE_SEPARATOR = re.compile(r"^(?:-{5,}Page\d+-{5,}|={5,}Page\d+={5,}|\*{5,}Entry\d+\*{5,})$", re.I)
MULTILINE_MARKERS = {"Message", "StringPicture", "ScrollText"}

QUOTE_PAIRS = OrderedDict([
    ("「」", ("「", "」")),
    ("“”", ("“", "”")),
    ("『』", ("『", "』")),
    ("‘’", ("‘", "’")),
    ('""', ('"', '"')),
    ("''", ("'", "'")),
    ("｢｣", ("｢", "｣")),
])
OPEN_TO_CLOSE = {op: cl for op, cl in QUOTE_PAIRS.values()}
ALL_QUOTES = set(OPEN_TO_CLOSE) | set(OPEN_TO_CLOSE.values())
TERMINALS = {"。", "；", ";", "！", "!", "？", "?"}
ELLIPSIS_RE = re.compile(r"(?:(?:[\.．·・。]\s*){2,}|…{1,}|……+)")

ORIGINAL_HEADERS = {"原文", "original"}
TRANSLATION_HEADERS = {"译文", "translation", "translated"}
CATEGORY_HEADERS = {"类别", "category", "type"}

LogFn = Callable[[str], None]


def noop(_: str) -> None:
    pass


@dataclass(frozen=True)
class ScopeItem:
    file_key: str
    group: str  # map/database/common_event/other
    record_count: int


@dataclass(frozen=True)
class SearchOptions:
    field: str = "both"  # original/translated/both
    mode: str = "keyword"  # keyword/exact
    case_sensitive: bool = False
    width_sensitive: bool = False
    strip_symbols: bool = False
    ignore_rpg_controls: bool = True


@dataclass(frozen=True)
class SearchResult:
    record: DataRecord
    matched_field: str


@dataclass(frozen=True)
class WidthIssue:
    record: DataRecord
    line_no: int
    width: float
    limit: float
    face_type: str
    visible_line: str


@dataclass(frozen=True)
class SpeakerCandidate:
    original_name: str
    translated_names: tuple[str, ...]
    pattern_type: int
    confidence: str
    occurrences: int
    records: tuple[DataRecord, ...]


@dataclass(frozen=True)
class QuoteIssue:
    record: DataRecord
    category: str  # 左引号/左右引号/手动
    reason: str
    auto_eligible: bool
    source_open: str | None
    source_close: str | None


@dataclass(frozen=True)
class PunctuationIssue:
    record: DataRecord
    reason: str
    proposed: str


@dataclass(frozen=True)
class TextChange:
    record: DataRecord
    reason: str
    proposed: str


@dataclass
class EditableDictionaryRow:
    original: str
    translation: str
    category: str = "其他"
    source: str = ""


@dataclass(frozen=True)
class DictionaryConflict:
    original: str
    translations: tuple[str, ...]


@dataclass(frozen=True)
class DatabaseProposal:
    record: DataRecord
    status: str  # 可翻译/新增/原文有改动/已一致/无默认译文
    baseline_original: str
    proposed_translation: str
    reason: str


# ---------- encoding-aware TXT parsing ----------

def _codec_name(value: str) -> str:
    value = (value or "936").strip().lower()
    aliases = {
        "932": "cp932", "cp932": "cp932", "shift-jis": "cp932", "shift_jis": "cp932",
        "936": "gbk", "cp936": "gbk", "gbk": "gbk",
        "utf-8": "utf-8", "utf8": "utf-8", "自动": "auto", "auto": "auto",
    }
    return aliases.get(value, value)


def smart_decode(raw: bytes, preferred: str = "936") -> tuple[str, str, bool, str]:
    bom = raw.startswith(b"\xef\xbb\xbf")
    payload = raw[3:] if bom else raw
    if bom:
        text = payload.decode("utf-8", errors="strict")
        return normalize_newlines(text), "utf-8", True, "\r\n" if b"\r\n" in raw else "\n"
    preferred_codec = _codec_name(preferred)
    candidates: list[str] = []
    # WindyTranslator 输出通常是 UTF-8。有效 UTF-8 且含脚本标记时优先，避免 936 默认导致乱码。
    try:
        utf = payload.decode("utf-8", errors="strict")
        if re.search(r"(?m)^\s*#[^#]+#", utf) or "Select Face Graphic:" in utf:
            return normalize_newlines(utf), "utf-8", False, "\r\n" if b"\r\n" in raw else "\n"
        candidates.append("utf-8")
    except UnicodeDecodeError:
        pass
    if preferred_codec != "auto":
        candidates.insert(0, preferred_codec)
    candidates.extend(["utf-8", "cp932", "gbk"])
    seen: set[str] = set()
    for codec in candidates:
        if codec in seen:
            continue
        seen.add(codec)
        try:
            text = payload.decode(codec, errors="strict")
            return normalize_newlines(text), codec, False, "\r\n" if b"\r\n" in raw else "\n"
        except (UnicodeDecodeError, LookupError):
            continue
    text = payload.decode(preferred_codec if preferred_codec != "auto" else "utf-8", errors="replace")
    return normalize_newlines(text), preferred_codec, False, "\r\n" if b"\r\n" in raw else "\n"


def read_text_file(path: Path, preferred: str = "936") -> tuple[str, str, bool, str]:
    return smart_decode(path.read_bytes(), preferred)


def write_text_file(path: Path, text: str, codec: str, bom: bool, newline: str) -> None:
    text = normalize_newlines(text)
    if newline != "\n":
        text = text.replace("\n", newline)
    payload = text.encode(codec, errors="strict")
    if bom and codec == "utf-8":
        payload = b"\xef\xbb\xbf" + payload
    path.write_bytes(payload)


def _parse_face(details: str) -> str:
    cleaned = details.strip().split("}", 1)[0].strip()
    if cleaned.lower() == "erase":
        return "NARRATION"
    parts = [x.strip() for x in cleaned.split(",")]
    name = (parts[0] if parts else "").strip("'\"")
    if not name or name.lower() == "erase":
        return "NARRATION"
    if len(parts) > 1 and parts[1].isdigit():
        return f"{name}_{parts[1]}"
    return name


def parse_string_script_text(text: str) -> list[ParsedOccurrence]:
    lines = normalize_newlines(text).splitlines(keepends=True)
    results: list[ParsedOccurrence] = []
    speaker = "NARRATION"
    counts: Counter[str] = Counter()
    i = 0
    while i < len(lines):
        stripped = lines[i].strip()
        if RE_PAGE_SEPARATOR.match(stripped):
            speaker = "NARRATION"; i += 1; continue
        face = RE_FACE.match(stripped)
        if face:
            speaker = _parse_face(face.group(1)); i += 1; continue
        mm = RE_MARKER_LINE.match(stripped)
        if not mm:
            i += 1; continue
        marker = mm.group(1)
        counts[marker] += 1
        inst = counts[marker]
        current_speaker = speaker if marker == "Message" else "SYSTEM"
        i += 1
        if marker in MULTILINE_MARKERS or marker.startswith("PluginCommand_"):
            start = i; block: list[str] = []
            while i < len(lines) and lines[i].strip() != "##":
                block.append(lines[i]); i += 1
            results.append(ParsedOccurrence(marker, current_speaker, "".join(block).rstrip("\n"), start, i, f"{marker}:{inst}:1", "multiline"))
            if i < len(lines): i += 1
            continue
        if marker == "EventName":
            if i < len(lines): i += 1
            continue
        if marker == "Choice":
            sub = 0
            while i < len(lines) and lines[i].strip() != "##":
                sub += 1
                results.append(ParsedOccurrence(marker, "SYSTEM", lines[i].rstrip("\n").strip(), i, i + 1, f"{marker}:{inst}:{sub}", "choice"))
                i += 1
            if i < len(lines): i += 1
            continue
        if i < len(lines):
            results.append(ParsedOccurrence(marker, "SYSTEM", lines[i].rstrip("\n").strip(), i, i + 1, f"{marker}:{inst}:1", "single"))
            i += 1
    return results


def parse_string_script(path: Path, preferred: str = "936") -> tuple[list[ParsedOccurrence], dict[str, object]]:
    text, codec, bom, newline = read_text_file(path, preferred)
    return parse_string_script_text(text), {"codec": codec, "bom": bom, "newline": newline}


def _list_txt(root: Path) -> OrderedDict[str, Path]:
    if not root.is_dir():
        raise QAError(f"TXT 文件夹不存在：{root}")
    out: OrderedDict[str, Path] = OrderedDict()
    for path in sorted(root.rglob("*.txt"), key=lambda p: str(p).casefold()):
        if BACKUP_DIRNAME in path.parts:
            continue
        out[str(PureWindowsPath(*path.relative_to(root).parts))] = path
    return out


def _line_loc(path: Path, occ: ParsedOccurrence | None) -> str:
    if occ is None:
        return f"{path}（未找到对应条目）"
    return f"{path}（第 {occ.start_line + 1}–{max(occ.start_line + 1, occ.end_line)} 行；{occ.locator}）"


def load_txt_records(
    origin_dir: Path,
    translated_dir: Path,
    origin_encoding: str = "936",
    translated_encoding: str = "936",
    log: LogFn = noop,
) -> SourceLoadResult:
    origins = _list_txt(origin_dir); translated = _list_txt(translated_dir)
    records: list[DataRecord] = []; warnings: list[str] = []
    for file_no, (file_key, opath) in enumerate(origins.items(), 1):
        log(f"读取 TXT [{file_no}/{len(origins)}]：{file_key}")
        oocc, ometa = parse_string_script(opath, origin_encoding)
        tpath = translated.get(file_key, translated_dir / Path(*PureWindowsPath(file_key).parts))
        if tpath.exists():
            tocc, tmeta = parse_string_script(tpath, translated_encoding)
        else:
            tocc, tmeta = [], dict(ometa)
            warnings.append(f"译文 TXT 缺失，将以原文作为当前译文：{file_key}")
        by_loc = {x.locator: x for x in tocc}
        for seq, occ in enumerate(oocc, 1):
            if occ.text == "": continue
            mt = by_loc.get(occ.locator)
            records.append(DataRecord(
                uid=f"txt|{file_key}|{occ.locator}", source_kind="txt", file_key=file_key,
                original=occ.text, translated=mt.text if mt else occ.text,
                marker=occ.marker, speaker_id=occ.speaker_id,
                original_location=_line_loc(opath, occ), translated_location=_line_loc(tpath, mt),
                original_open_path=opath, translated_open_path=tpath,
                update_ref=(tpath, occ.locator, occ.kind),
                metadata={"original_occurrence": occ, "translated_occurrence": mt, "origin_format": ometa,
                          "translated_format": tmeta, "origin_dir": origin_dir, "translated_dir": translated_dir,
                          "sequence": seq},
            ))
    for extra in sorted(set(translated) - set(origins)):
        warnings.append(f"译文文件夹中存在原文文件夹没有的 TXT，已忽略：{extra}")
    return SourceLoadResult(tuple(records), tuple(warnings), "txt")




def _iter_excel_workbooks(table_dir: Path) -> Iterator[Path]:
    if not table_dir.is_dir():
        raise QAError(f"Excel 文件夹不存在：{table_dir}")
    for suffix in ("*.xlsx", "*.xlsm"):
        for path in sorted(table_dir.rglob(suffix), key=lambda p: str(p).casefold()):
            if BACKUP_DIRNAME in path.parts or path.name.startswith(("~$", ".")):
                continue
            yield path


def _canonical_excel_file_key(relative: str) -> str:
    p = PureWindowsPath(relative.replace("/", "\\"))
    if p.suffix.lower() in {".xlsx", ".xlsm"}:
        p = p.with_suffix(".txt")
    return str(p)


def load_excel_records(table_dir: Path, log: LogFn = noop) -> SourceLoadResult:
    workbooks = list(_iter_excel_workbooks(table_dir))
    if not workbooks:
        raise QAError(f"Excel 文件夹中没有 .xlsx/.xlsm 文件：{table_dir}")
    records: list[DataRecord] = []
    warnings: list[str] = []
    for idx, path in enumerate(workbooks, 1):
        relative = str(path.relative_to(table_dir))
        log(f"读取 Excel [{idx}/{len(workbooks)}]：{relative}")
        try:
            wb = load_workbook(path, read_only=True, data_only=False)
        except Exception as exc:
            warnings.append(f"无法读取，已跳过：{relative}（{exc}）")
            continue
        try:
            sheets = [wb[TRANSLATION_SHEET]] if TRANSLATION_SHEET in wb.sheetnames else list(wb.worksheets)
            found = False
            for ws in sheets:
                max_row, _ = _worksheet_dimensions(ws)
                if max_row < 1:
                    continue
                first = next(ws.iter_rows(min_row=1, max_row=1))
                values = [c.value for c in first]
                headers = {str(v).strip(): i for i, v in enumerate(values, 1) if v is not None}
                ocol = _find_header(headers, ORIGINAL_HEADERS)
                tcol = _find_header(headers, TRANSLATION_HEADERS)
                if ocol is None or tcol is None:
                    continue
                mcol = _find_header(headers, {"标记", "original_marker", "marker"})
                scol = _find_header(headers, {"说话人ID", "speaker_id", "speaker"})
                pcol = _find_header(headers, {"相对TXT路径", "相对txt路径", "file_key", "filepath"})
                found = True
                for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                    original = row[ocol-1] if len(row) >= ocol else None
                    if original in (None, ""):
                        continue
                    translated = row[tcol-1] if len(row) >= tcol else None
                    marker = row[mcol-1] if mcol and len(row) >= mcol else ""
                    speaker = row[scol-1] if scol and len(row) >= scol else ""
                    logical = row[pcol-1] if pcol and len(row) >= pcol else None
                    file_key = str(logical) if logical not in (None, "") else _canonical_excel_file_key(relative)
                    base = f"{path} / {ws.title}"
                    uid = f"excel|{relative}|{ws.title}|{row_no}"
                    records.append(DataRecord(
                        uid=uid, source_kind="excel", file_key=file_key, original=str(original),
                        translated=str(original if translated is None else translated), marker=str(marker or ""), speaker_id=str(speaker or ""),
                        original_location=f"{base}!{_column_letter(ocol)}{row_no}", translated_location=f"{base}!{_column_letter(tcol)}{row_no}",
                        original_open_path=path, translated_open_path=path, update_ref=(path, ws.title, row_no, tcol),
                        metadata={"table_dir": table_dir, "workbook_relative": relative, "sequence": row_no-1},
                    ))
                break
            if not found:
                warnings.append(f"未找到原文/译文列，已跳过：{relative}")
        finally:
            wb.close()
    if not records:
        raise QAError("没有找到可分析的 Excel 翻译记录。")
    return SourceLoadResult(tuple(records), tuple(warnings), "excel")


# ---------- source scan and scopes ----------

def classify_file_key(file_key: str) -> str:
    normalized = str(file_key).replace("/", "\\")
    name = PureWindowsPath(normalized).name
    if MAP_RE.search(normalized) or re.fullmatch(r"Map\d+\.xlsx?", name, re.I):
        return "map"
    if COMMON_EVENT_RE.search(normalized):
        return "common_event"
    if "database" in normalized.casefold() or not re.match(r"Map\d+", name, re.I):
        return "database"
    return "other"


def build_scope_items(records: Sequence[DataRecord]) -> list[ScopeItem]:
    counts: OrderedDict[str, int] = OrderedDict()
    for rec in records:
        counts[rec.file_key] = counts.get(rec.file_key, 0) + 1
    return [ScopeItem(k, classify_file_key(k), v) for k, v in counts.items()]


def filter_records(records: Sequence[DataRecord], selected_file_keys: set[str], workset: set[str] | None = None) -> list[DataRecord]:
    out = [r for r in records if r.file_key in selected_file_keys]
    if workset:
        out = [r for r in out if r.uid in workset]
    return out


# ---------- text utilities ----------

def visible_text(text: str) -> str:
    return strip_control_codes(normalize_newlines(text))


def normalize_search(text: str, *, case_sensitive: bool, width_sensitive: bool, strip_symbols: bool) -> str:
    value = visible_text(text)
    if not width_sensitive:
        value = unicodedata.normalize("NFKC", value)
    if strip_symbols:
        value = "".join(ch for ch in value if ch.isalnum() or unicodedata.category(ch).startswith("L"))
    else:
        # Preserve internal spaces and line breaks so exact multi-line matching is truly exact.
        value = normalize_newlines(value).strip()
    if not case_sensitive:
        value = value.casefold()
    return value


def search_records(records: Iterable[DataRecord], query: str, options: SearchOptions) -> list[SearchResult]:
    needle = normalize_search(query, case_sensitive=options.case_sensitive, width_sensitive=options.width_sensitive, strip_symbols=options.strip_symbols)
    if not needle:
        raise QAError("查询内容在去除控制符和特殊符号后为空。")
    out: list[SearchResult] = []
    for r in records:
        fields = []
        if options.field in {"original", "both"}: fields.append(("原文", r.original))
        if options.field in {"translated", "both"}: fields.append(("译文", r.translated))
        matched: list[str] = []
        for label, value in fields:
            hay = normalize_search(value, case_sensitive=options.case_sensitive, width_sensitive=options.width_sensitive, strip_symbols=options.strip_symbols)
            ok = needle in hay if options.mode == "keyword" else needle == hay
            if ok: matched.append(label)
        if matched:
            out.append(SearchResult(r, "＋".join(matched)))
    return out


def is_face_message(record: DataRecord) -> bool:
    return record.marker == "Message" and (record.speaker_id or "").strip().upper() not in NON_FACE_IDS


def is_narration_message(record: DataRecord) -> bool:
    return record.marker == "Message" and (record.speaker_id or "").strip().upper() in {"", "NARRATION", "_ERASE_FACE_"}


def analyze_width(records: Iterable[DataRecord], face_limit: float, narration_limit: float, check_face: bool, check_narration: bool) -> list[WidthIssue]:
    issues: list[WidthIssue] = []
    for r in records:
        if r.marker != "Message": continue
        if is_face_message(r):
            if not check_face: continue
            limit, face_type = face_limit, "有头像"
        elif is_narration_message(r):
            if not check_narration: continue
            limit, face_type = narration_limit, "无头像"
        else:
            continue
        for i, line in enumerate(normalize_newlines(r.translated).split("\n"), 1):
            vis = strip_control_codes(line)
            width = fullwidth_units(vis)
            if width > limit:
                issues.append(WidthIssue(r, i, width, limit, face_type, vis))
    return issues


def records_context(records: Sequence[DataRecord], record: DataRecord, radius: int = 3) -> list[DataRecord]:
    same = [r for r in records if r.file_key == record.file_key]
    try: idx = next(i for i, r in enumerate(same) if r.uid == record.uid)
    except StopIteration: return []
    return same[max(0, idx-radius):idx+radius+1]


# ---------- generic writes ----------

def _timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _backup_root(base: Path, prefix: str = "QA") -> Path:
    root = base / BACKUP_DIRNAME / f"{prefix}_{_timestamp()}"
    n = 1
    while root.exists():
        root = base / BACKUP_DIRNAME / f"{prefix}_{_timestamp()}_{n}"; n += 1
    return root


def _atomic_json(path: Path, data: Mapping) -> None:
    temp = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
    with temp.open("w", encoding="utf-8", newline="\n") as fh:
        json.dump(data, fh, ensure_ascii=False, indent=4)
        fh.flush(); os.fsync(fh.fileno())
    os.replace(temp, path)


def _format_occurrence(lines: list[str], occ: ParsedOccurrence, translated: str) -> list[str]:
    translated = normalize_newlines(translated)
    if occ.kind == "multiline":
        original_block = "".join(lines[occ.start_line:occ.end_line])
        replacement = translated
        if original_block.endswith("\n") and not replacement.endswith("\n"): replacement += "\n"
        elif not original_block.endswith("\n") and replacement.endswith("\n"): replacement = replacement.rstrip("\n")
        return replacement.splitlines(keepends=True) or ([replacement] if replacement else [])
    raw = lines[occ.start_line].rstrip("\n")
    lead = raw[:len(raw)-len(raw.lstrip())]
    trail_len = len(raw)-len(raw.rstrip()); trail = raw[len(raw)-trail_len:] if trail_len else ""
    return [f"{lead}{translated.strip()}{trail}\n"]


def apply_translation_updates(updates: Mapping[str, tuple[DataRecord, str]], log: LogFn = noop) -> tuple[int, Path]:
    if not updates: raise QAError("没有需要保存的修改。")
    kinds = {r.source_kind for r, _ in updates.values()}
    if len(kinds) != 1: raise QAError("一次保存只能处理同一种数据源。")
    kind = next(iter(kinds)); records = [r for r, _ in updates.values()]
    if kind == "json":
        path = records[0].translated_open_path
        root = _backup_root(path.parent, "Edit"); root.mkdir(parents=True, exist_ok=True); shutil.copy2(path, root/path.name)
        with path.open("r", encoding="utf-8-sig") as fh: data = json.load(fh, object_pairs_hook=_no_duplicate_object)
        for rec, proposed in updates.values():
            file_key, original_key = rec.update_ref
            if data[file_key][original_key]["text"] != rec.translated:
                raise QAError(f"JSON 在扫描后已改变：{file_key} / {original_key[:30]}")
            data[file_key][original_key]["text"] = proposed
        _atomic_json(path, data); return len(updates), root
    if kind == "excel":
        table_dir = Path(str(records[0].metadata.get("table_dir") or records[0].translated_open_path.parent))
        root = _backup_root(table_dir, "Edit")
        grouped: dict[Path, list[tuple[DataRecord, str]]] = defaultdict(list)
        for rec, proposed in updates.values(): grouped[rec.translated_open_path].append((rec, proposed))
        count = 0
        for path, items in grouped.items():
            rel = path.relative_to(table_dir) if path.is_relative_to(table_dir) else Path(path.name)
            bp = root/rel; bp.parent.mkdir(parents=True, exist_ok=True); shutil.copy2(path, bp)
            wb = load_workbook(path, read_only=False, data_only=False, keep_vba=path.suffix.lower()==".xlsm")
            try:
                for rec, proposed in items:
                    _, sheet, row, col = rec.update_ref; cell = wb[sheet].cell(row, col)
                    current = "" if cell.value is None else str(cell.value)
                    if current != rec.translated: raise QAError(f"Excel 在扫描后已改变：{path} / {sheet}!{cell.coordinate}")
                    cell.value = proposed; cell.data_type = "s"; count += 1
                temp = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp{path.suffix}"); wb.save(temp)
            finally: wb.close()
            os.replace(temp, path)
        return count, root
    if kind == "txt":
        translated_dir = Path(str(records[0].metadata.get("translated_dir") or records[0].translated_open_path.parent))
        root = _backup_root(translated_dir.parent, "Edit")
        grouped: dict[Path, list[tuple[DataRecord, str]]] = defaultdict(list)
        for rec, proposed in updates.values(): grouped[rec.translated_open_path].append((rec, proposed))
        count = 0
        for path, items in grouped.items():
            if not path.exists():
                path.parent.mkdir(parents=True, exist_ok=True); shutil.copy2(items[0][0].original_open_path, path)
            rel = path.relative_to(translated_dir) if path.is_relative_to(translated_dir) else Path(path.name)
            bp = root/translated_dir.name/rel; bp.parent.mkdir(parents=True, exist_ok=True); shutil.copy2(path, bp)
            fmt = items[0][0].metadata.get("translated_format") or {"codec":"utf-8","bom":False,"newline":"\n"}
            text, codec, bom, newline = read_text_file(path, str(fmt.get("codec", "utf-8")))
            occs = {x.locator:x for x in parse_string_script_text(text)}
            lines = text.splitlines(keepends=True); replacements=[]
            for rec, proposed in items:
                _, locator, _kind = rec.update_ref; occ = occs.get(locator)
                if occ is None: raise QAError(f"TXT 找不到条目：{path} / {locator}")
                if occ.text != rec.translated: raise QAError(f"TXT 在扫描后已改变：{path} / {locator}")
                replacements.append((occ.start_line, occ.end_line, _format_occurrence(lines, occ, proposed))); count += 1
            for start, end, repl in sorted(replacements, reverse=True): lines[start:end] = repl
            temp = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
            write_text_file(temp, "".join(lines), codec, bom, newline); os.replace(temp, path)
        return count, root
    raise QAError(f"不支持的数据源：{kind}")


# ---------- speaker analysis ----------

def _clean_line(line: str) -> str:
    return strip_control_codes(line).strip()


def _first_nonempty(lines: list[str]) -> tuple[int, str] | None:
    for i, line in enumerate(lines):
        clean = _clean_line(line)
        if clean: return i, clean
    return None


def _contains_open_quote(text: str) -> bool:
    return any(q in text for q in OPEN_TO_CLOSE)


def detect_speaker(text: str, pattern_mode: int = 0) -> tuple[str, int, str] | None:
    lines = normalize_newlines(text).split("\n")[:4]
    first = _first_nonempty(lines)
    if not first: return None
    idx, line = first
    later = [_clean_line(x) for x in lines[idx+1:4]]
    # Common RPG nameplate style: a whole first line wrapped in 【】 or similar brackets.
    m = re.match(r"^[【〔［\[]\s*(.+?)\s*[】〕］\]]$", line)
    if m and pattern_mode in {0, 1}:
        return m.group(1).strip(), 1, "确定"
    # type 3: single line ending colon
    m = re.match(r"^(.+?)[：:]\s*$", line)
    if m and pattern_mode in {0, 3}:
        return m.group(1).strip(), 3, "确定"
    # type 4: same line separated by colon or quote
    m = re.match(r"^([^「『“‘\"']+?)[：:]\s*(.+)$", line)
    if m and pattern_mode in {0, 4}:
        return m.group(1).strip(), 4, "较确定"
    quote_positions = [(line.find(q), q) for q in OPEN_TO_CLOSE if q in line]
    quote_positions = [x for x in quote_positions if x[0] >= 0]
    if quote_positions:
        pos, _q = min(quote_positions)
        prefix = re.sub(r"[：:]\s*$", "", line[:pos]).strip()
        if prefix and pattern_mode in {0, 4}:
            return prefix, 4, "需检查"
    # type1: speaker line, following starts quote
    if pattern_mode in {0, 1} and not _contains_open_quote(line) and any(_contains_open_quote(x) for x in later):
        return re.sub(r"[：:]\s*$", "", line).strip(), 1, "较确定" if pattern_mode == 1 else "需检查"
    # type2 only manual
    if pattern_mode == 2 and not _contains_open_quote(line) and any(x for x in later):
        return line.strip(), 2, "手动规则"
    return None


def analyze_speakers(records: Sequence[DataRecord], pattern_mode: int = 0, first_n: int | None = None) -> list[SpeakerCandidate]:
    selected = [r for r in records if r.marker == "Message"]
    if first_n is not None: selected = selected[:max(0, first_n)]
    grouped: dict[str, dict[str, object]] = {}
    for r in selected:
        found = detect_speaker(r.original, pattern_mode)
        if not found: continue
        name, ptype, confidence = found
        tfound = detect_speaker(r.translated, ptype)
        tname = tfound[0] if tfound else ""
        g = grouped.setdefault(name, {"translations":set(), "ptype":ptype, "confidence":confidence, "records":[]})
        if tname: g["translations"].add(tname)
        g["records"].append(r)
    out=[]
    for name, g in grouped.items():
        out.append(SpeakerCandidate(name, tuple(sorted(g["translations"])), int(g["ptype"]), str(g["confidence"]), len(g["records"]), tuple(g["records"])))
    return sorted(out, key=lambda x:(-x.occurrences, x.original_name))


# ---------- quotes ----------

def _mask_controls(text: str) -> list[bool]:
    mask=[False]*len(text)
    for m in CONTROL_CODE_RE.finditer(text):
        for i in range(m.start(), m.end()): mask[i]=True
    return mask


def _visible_positions(text: str) -> list[int]:
    mask=_mask_controls(text); return [i for i,ch in enumerate(text) if not mask[i] and not ch.isspace()]


def _line_start_positions(text: str) -> set[int]:
    mask=_mask_controls(text); starts=set(); at=True
    for i,ch in enumerate(text):
        if ch=="\n": at=True; continue
        if mask[i] or ch.isspace(): continue
        if at: starts.add(i)
        at=False
    return starts


def _quote_tokens(text: str) -> list[tuple[int,str]]:
    mask=_mask_controls(text); return [(i,ch) for i,ch in enumerate(text) if not mask[i] and ch in ALL_QUOTES]


def classify_quote_issue(record: DataRecord) -> QuoteIssue | None:
    if record.marker != "Message": return None
    source=record.original; trans=record.translated
    positions=_visible_positions(source)
    if not positions: return None
    last=positions[-1]; line_starts=_line_start_positions(source); tokens=_quote_tokens(source)
    # Explicitly safe: matching pair with close at text end
    for op, cl in OPEN_TO_CLOSE.items():
        opens=[p for p,c in tokens if c==op]; closes=[p for p,c in tokens if c==cl]
        if op==cl:
            if len(opens)>=2 and opens[-1]==last:
                start=opens[0]
                return _compare_quote_target(record, "左右引号", op, cl, start in line_starts)
        elif opens and closes and closes[-1]==last:
            return _compare_quote_target(record, "左右引号", op, cl, opens[0] in line_starts)
    # Safe left-only: line-start open without corresponding close
    for op, cl in OPEN_TO_CLOSE.items():
        opens=[p for p,c in tokens if c==op and p in line_starts]
        closes=[p for p,c in tokens if c==cl]
        if opens and (not closes or (op==cl and len(opens)%2==1)):
            return _compare_quote_target(record, "只有左引号", op, None, True)
    # Anything with quote but not safe => manual
    if tokens:
        return QuoteIssue(record, "手动", "引号位于文本内部或配对位置不明确，不能批量处理", False, None, None)
    return None


def _compare_quote_target(record: DataRecord, category: str, op: str, cl: str | None, at_line_start: bool) -> QuoteIssue | None:
    ttokens=_quote_tokens(record.translated); tvis=_visible_positions(record.translated)
    if not tvis: return QuoteIssue(record, category, "译文为空或没有可见文本", False, op, cl)
    starts=_line_start_positions(record.translated); last=tvis[-1]
    if category=="左右引号":
        ok_open=any(c==op and p in starts for p,c in ttokens) if at_line_start else any(c==op for p,c in ttokens)
        ok_close=any(c==cl and p==last for p,c in ttokens)
        if ok_open and ok_close: return None
        return QuoteIssue(record, category, "译文缺少相应引号或引号样式与原文不一致", True, op, cl)
    ok_open=any(c==op and p in starts for p,c in ttokens)
    extra_close=any(c in set(OPEN_TO_CLOSE.values()) and p==last for p,c in ttokens)
    if ok_open and not extra_close: return None
    return QuoteIssue(record, category, "译文左引号缺失/样式不一致，或存在不应有的右引号", True, op, None)


def analyze_quotes(records: Iterable[DataRecord]) -> list[QuoteIssue]:
    return [x for r in records if (x:=classify_quote_issue(r)) is not None]


def _remove_edge_quotes(text: str) -> str:
    chars=list(text); mask=_mask_controls(text); starts=_line_start_positions(text); vis=_visible_positions(text)
    remove=set()
    for i,ch in enumerate(chars):
        if not mask[i] and i in starts and ch in ALL_QUOTES: remove.add(i)
    if vis and chars[vis[-1]] in ALL_QUOTES: remove.add(vis[-1])
    return "".join(ch for i,ch in enumerate(chars) if i not in remove)


def build_quote_proposal(record: DataRecord, form: str, pair: tuple[str,str]) -> str:
    text=_remove_edge_quotes(record.translated)
    op,cl=pair
    if form=="删除引号": return text
    # Insert open at the first visible character of the first nonempty line; controls remain before it.
    vis=_visible_positions(text)
    if not vis: return text
    pos=vis[0]; text=text[:pos]+op+text[pos:]
    if form=="左右引号":
        vis2=_visible_positions(text); end=vis2[-1]+1 if vis2 else len(text); text=text[:end]+cl+text[end:]
    return text


# ---------- punctuation ----------

def _terminal_info(text: str) -> tuple[str|None, int|None]:
    mask=_mask_controls(text); vis=[i for i,ch in enumerate(text) if not mask[i] and not ch.isspace()]
    while vis and text[vis[-1]] in ALL_QUOTES: vis.pop()
    if not vis: return None,None
    pos=vis[-1]; ch=text[pos]
    return (ch,pos) if ch in TERMINALS else (None,pos)


def _set_terminal(text: str, target: str | None, remove_english_period: bool=True) -> str:
    ch,pos=_terminal_info(text)
    if pos is None: return text
    chars=list(text)
    if ch in TERMINALS:
        if target is None: del chars[pos]
        else: chars[pos]=target
    elif target:
        # remove a terminal English/fullwidth period before adding a valid terminal.
        if remove_english_period and chars[pos] in {".","．"}: chars[pos]=target
        else: chars.insert(pos+1,target)
    return "".join(chars)


def analyze_punctuation(records: Iterable[DataRecord], mode: str="source", default_mark: str="。", delete_marks: set[str]|None=None) -> list[PunctuationIssue]:
    out=[]; delete_marks=delete_marks or set(TERMINALS)
    for r in records:
        if r.marker!="Message": continue
        so,_=_terminal_info(r.original); st,_=_terminal_info(r.translated)
        proposed=r.translated; reason=""
        if mode=="source":
            if so != st:
                proposed=_set_terminal(r.translated, so); reason=f"按原文同步句尾符号：{st or '无'} → {so or '无'}"
        elif mode=="missing":
            if st is None:
                proposed=_set_terminal(r.translated, so or default_mark); reason=f"补齐句尾符号：{so or default_mark}"
        elif mode=="delete":
            if st in delete_marks:
                proposed=_set_terminal(r.translated, None); reason=f"删除句尾符号：{st}"
        if proposed!=r.translated: out.append(PunctuationIssue(r,reason,proposed))
    return out


# ---------- replacements and transformations ----------

def _outside_controls_transform(text: str, fn: Callable[[str],str]) -> str:
    out=[]; last=0
    for m in CONTROL_CODE_RE.finditer(text):
        out.append(fn(text[last:m.start()])); out.append(m.group(0)); last=m.end()
    out.append(fn(text[last:])); return "".join(out)


def replace_outside_controls(text: str, old: str, new: str) -> str:
    if not old: return text
    return _outside_controls_transform(text, lambda s:s.replace(old,new))


def analyze_symbol_counts(records: Iterable[DataRecord]) -> list[tuple[str,int,int]]:
    oc=Counter(); tc=Counter()
    for r in records:
        for ch in visible_text(r.original):
            if unicodedata.category(ch).startswith("P") or ch in ALL_QUOTES: oc[ch]+=1
        for ch in visible_text(r.translated):
            if unicodedata.category(ch).startswith("P") or ch in ALL_QUOTES: tc[ch]+=1
    return [(ch,oc[ch],tc[ch]) for ch in sorted(set(oc)|set(tc), key=lambda x:(-tc[x]-oc[x],x))]


def build_symbol_changes(records: Iterable[DataRecord], old: str, new: str) -> list[TextChange]:
    out=[]
    for r in records:
        proposed=replace_outside_controls(r.translated,old,new)
        if proposed!=r.translated: out.append(TextChange(r,f"符号替换：{old} → {new}",proposed))
    return out


def to_fullwidth_ascii(s: str) -> str:
    return "".join(chr(ord(ch)+0xFEE0) if 0x21<=ord(ch)<=0x7E else ("　" if ch==" " else ch) for ch in s)


def to_halfwidth_ascii(s: str) -> str:
    return "".join(chr(ord(ch)-0xFEE0) if 0xFF01<=ord(ch)<=0xFF5E else (" " if ch=="　" else ch) for ch in s)


def transform_alnum_segment(s: str, mode: str) -> str:
    if mode=="全角": return to_fullwidth_ascii(s)
    if mode=="半角": return to_halfwidth_ascii(s)
    if mode=="大写": return re.sub(r"[A-Za-z]+", lambda m:m.group(0).upper(), s)
    if mode=="小写": return re.sub(r"[A-Za-z]+", lambda m:m.group(0).lower(), s)
    if mode=="首字母大写": return re.sub(r"[A-Za-z]+", lambda m:m.group(0)[:1].upper()+m.group(0)[1:].lower(), s)
    return s


def build_alnum_changes(records: Iterable[DataRecord], mode: str, term: str="", case_sensitive: bool=False) -> list[TextChange]:
    out=[]
    for r in records:
        if term:
            flags=0 if case_sensitive else re.I
            pattern=re.compile(re.escape(term),flags)
            proposed=_outside_controls_transform(r.translated, lambda s:pattern.sub(lambda m:transform_alnum_segment(m.group(0),mode),s))
        else:
            proposed=_outside_controls_transform(r.translated, lambda s:transform_alnum_segment(s,mode))
        if proposed!=r.translated: out.append(TextChange(r,f"英语数字格式：{mode}"+(f"（仅 {term}）" if term else ""),proposed))
    return out


def analyze_ellipsis(records: Iterable[DataRecord], replacement: str="……") -> list[TextChange]:
    out=[]
    for r in records:
        proposed=_outside_controls_transform(r.translated, lambda s:ELLIPSIS_RE.sub(replacement,s))
        if proposed!=r.translated: out.append(TextChange(r,f"省略号统一为 {replacement}",proposed))
    return out


# ---------- dictionaries ----------

def _decode_csv(path: Path) -> str:
    raw=path.read_bytes()
    for enc in ("utf-8-sig","gb18030","cp932"):
        try:return raw.decode(enc)
        except UnicodeDecodeError: pass
    return raw.decode("utf-8",errors="replace")


def _dict_columns(headers: Sequence[object]) -> tuple[int,int,int|None]:
    normalized={_normalize_header(v):i for i,v in enumerate(headers)}
    oi=next((normalized[x] for x in map(_normalize_header,ORIGINAL_HEADERS) if x in normalized),None)
    ti=next((normalized[x] for x in map(_normalize_header,TRANSLATION_HEADERS) if x in normalized),None)
    ci=next((normalized[x] for x in map(_normalize_header,CATEGORY_HEADERS) if x in normalized),None)
    if oi is None or ti is None: raise QAError("辞典必须包含 Original/Translation 或 原文/译文列。")
    return oi,ti,ci


def load_editable_dictionary(path: Path) -> list[EditableDictionaryRow]:
    rows=[]
    if path.suffix.lower()==".csv":
        data=list(csv.reader(_decode_csv(path).splitlines()))
    else:
        wb=load_workbook(path,read_only=True,data_only=True); ws=wb.active; data=list(ws.iter_rows(values_only=True)); wb.close()
    if not data:return []
    oi,ti,ci=_dict_columns(data[0])
    for row in data[1:]:
        if len(row)<=max(oi,ti):continue
        o=str(row[oi] or "").strip(); t=str(row[ti] or "").strip()
        if not o or not t:continue
        c=str(row[ci] or "其他").strip() if ci is not None and len(row)>ci else "其他"
        rows.append(EditableDictionaryRow(o,t,c,path.name))
    return rows


def dictionary_conflicts(rows: Sequence[EditableDictionaryRow]) -> list[DictionaryConflict]:
    groups: dict[str,set[str]]=defaultdict(set)
    for row in rows:
        if row.original and row.translation: groups[row.original].add(row.translation)
    return [DictionaryConflict(o,tuple(sorted(ts))) for o,ts in groups.items() if len(ts)>1]


def save_dictionary(path: Path, rows: Sequence[EditableDictionaryRow]) -> None:
    rows=sorted(rows,key=lambda x:(x.original,x.translation,x.category))
    if path.suffix.lower()==".csv":
        with path.open("w",encoding="utf-8-sig",newline="") as fh:
            w=csv.writer(fh); w.writerow(["Original","Translation","Category"])
            for r in rows:w.writerow([r.original,r.translation,r.category])
        return
    wb=Workbook(); ws=wb.active; ws.title="Dictionary"; ws.append(["Original","Translation","Category"])
    for r in rows:ws.append([r.original,r.translation,r.category])
    for cell in ws[1]:cell.font=Font(bold=True);cell.fill=PatternFill("solid",fgColor="D9EAF7")
    ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:C{max(1,ws.max_row)}"
    ws.column_dimensions["A"].width=34;ws.column_dimensions["B"].width=34;ws.column_dimensions["C"].width=16
    wb.save(path);wb.close()


def dictionary_entries_from_rows(rows: Sequence[EditableDictionaryRow], conflict_mode: str="first") -> tuple[list[DictionaryEntry], list[DictionaryConflict]]:
    conflicts=dictionary_conflicts(rows); chosen=[]; seen=set()
    for r in rows:
        if conflict_mode=="first" and r.original in seen:continue
        seen.add(r.original);chosen.append(DictionaryEntry(r.original,r.translation,Path(r.source or "内置辞典")))
    return chosen,conflicts


def analyze_dictionary(records: Iterable[DataRecord], entries: Sequence[DictionaryEntry], messages_only: bool=False) -> list[DictionaryWarning]:
    out=[]
    for r in records:
        if messages_only and r.marker!="Message":continue
        matches=longest_non_overlapping_matches(strip_control_codes(r.original),entries)
        for e in matches:
            if e.translation not in strip_control_codes(r.translated):
                out.append(DictionaryWarning(r,e.original,e.translation,e.source_file,strip_control_codes(r.original),strip_control_codes(r.translated)))
    return out


def build_used_dictionary(records: Iterable[DataRecord], rows: Sequence[EditableDictionaryRow]) -> list[EditableDictionaryRow]:
    # Preserve multiple translations, but longest-match original terms only.
    entries,_=dictionary_entries_from_rows(rows,"all")
    by_pair={(r.original,r.translation):r for r in rows}; used=set()
    for rec in records:
        for e in longest_non_overlapping_matches(strip_control_codes(rec.original),entries):
            used.add((e.original,e.translation))
    return [by_pair[p] for p in sorted(used)]


# ---------- database direct translation ----------

def database_records(records: Iterable[DataRecord]) -> list[DataRecord]:
    return [r for r in records if classify_file_key(r.file_key)=="database" and not COMMON_EVENT_RE.search(r.file_key)]


def compare_database(current: Sequence[DataRecord], baseline: Sequence[DataRecord]|None=None, dictionary_rows: Sequence[EditableDictionaryRow]|None=None) -> list[DatabaseProposal]:
    current=database_records(current); out=[]
    if dictionary_rows is not None:
        mapping=OrderedDict()
        for row in dictionary_rows: mapping.setdefault(row.original,row.translation)
        for r in current:
            if r.original in mapping:
                prop=mapping[r.original];status="已一致" if r.translated==prop else "可翻译"
                out.append(DatabaseProposal(r,status,r.original,prop,"按辞典原文精确匹配"))
            else:out.append(DatabaseProposal(r,"无默认译文","","","辞典中没有精确原文"))
        return out
    baseline=database_records(baseline or [])
    grouped: dict[str,list[DataRecord]]=defaultdict(list)
    for r in baseline:grouped[r.file_key].append(r)
    current_grouped: dict[str,list[DataRecord]]=defaultdict(list)
    for r in current:current_grouped[r.file_key].append(r)
    for file_key, rows in current_grouped.items():
        brow=grouped.get(file_key,[])
        for idx,r in enumerate(rows):
            if idx>=len(brow):out.append(DatabaseProposal(r,"新增","","","默认数据库中没有对应位置"));continue
            b=brow[idx]
            if b.original!=r.original:
                out.append(DatabaseProposal(r,"原文有改动",b.original,b.translated,"同一文件同一顺序位置的原文发生变化"))
            else:
                status="已一致" if r.translated==b.translated else "可翻译"
                out.append(DatabaseProposal(r,status,b.original,b.translated,"文件与顺序位置一致，原文相同"))
    return out


# ---------- config ----------

def save_config(path: Path, data: Mapping) -> None:
    temp=path.with_suffix(path.suffix+".tmp");temp.write_text(json.dumps(data,ensure_ascii=False,indent=2),encoding="utf-8");os.replace(temp,path)


def load_config(path: Path) -> dict:
    try:return json.loads(path.read_text(encoding="utf-8"))
    except Exception:return {}

# ============================================================================
# v4 additions: configurable display/search, font-aware wrapping, rich speaker
# aggregation, punctuation catalogue, ellipsis classifications and duplicates.
# ============================================================================

CONFIDENCE_RANK = {"手动规则": 0, "需检查": 1, "较确定": 2, "确定": 3}

@dataclass(frozen=True)
class SpeakerTranslationOption:
    translation: str
    count: int
    pattern_types: tuple[int, ...]
    confidence: str


@dataclass(frozen=True)
class SpeakerGroup:
    original_name: str
    options: tuple[SpeakerTranslationOption, ...]
    occurrences: int
    pattern_types: tuple[int, ...]
    max_confidence: str
    records: tuple[DataRecord, ...]


@dataclass(frozen=True)
class QuoteAnalysisRow:
    record: DataRecord
    mode: str  # 引号状况 / 引号样式
    status: str
    auto_eligible: bool
    reason: str
    source_styles: tuple[str, ...]
    translated_styles: tuple[str, ...]
    proposal: str | None = None


@dataclass(frozen=True)
class SymbolCatalogEntry:
    symbol: str
    category: str
    description: str
    peers: tuple[str, ...]

    @property
    def unicode_label(self) -> str:
        return " ".join(f"U+{ord(ch):04X}" for ch in self.symbol)


@dataclass(frozen=True)
class EllipsisOccurrence:
    record: DataRecord
    side: str
    raw: str
    visible_style: str
    kind: str
    point_char: str
    count: int
    interrupted_by_control: bool
    start: int
    end: int
    manual_only: bool = False
    reason: str = ""


@dataclass(frozen=True)
class DuplicateTextGroup:
    original: str
    translations: tuple[tuple[str, int], ...]
    records: tuple[DataRecord, ...]

    @property
    def consistent(self) -> bool:
        return len(self.translations) <= 1


def normalize_search(text: str, *, case_sensitive: bool, width_sensitive: bool,
                     strip_symbols: bool, ignore_rpg_controls: bool = True) -> str:
    value = normalize_newlines(text or "")
    if ignore_rpg_controls:
        value = strip_control_codes(value)
    if not width_sensitive:
        value = unicodedata.normalize("NFKC", value)
    if strip_symbols:
        value = "".join(ch for ch in value if ch.isalnum() or unicodedata.category(ch).startswith("L"))
    else:
        # Preserve internal spaces and line breaks so exact multi-line matching is truly exact.
        value = normalize_newlines(value).strip()
    if not case_sensitive:
        value = value.casefold()
    return value


def search_records(records: Iterable[DataRecord], query: str, options: SearchOptions) -> list[SearchResult]:
    needle = normalize_search(
        query,
        case_sensitive=options.case_sensitive,
        width_sensitive=options.width_sensitive,
        strip_symbols=options.strip_symbols,
        ignore_rpg_controls=options.ignore_rpg_controls,
    )
    if not needle:
        raise QAError("查询内容在应用当前查询规则后为空。")
    out: list[SearchResult] = []
    for r in records:
        fields: list[tuple[str, str]] = []
        if options.field in {"original", "both"}:
            fields.append(("原文", r.original))
        if options.field in {"translated", "both"}:
            fields.append(("译文", r.translated))
        matched: list[str] = []
        for label, value in fields:
            hay = normalize_search(
                value,
                case_sensitive=options.case_sensitive,
                width_sensitive=options.width_sensitive,
                strip_symbols=options.strip_symbols,
                ignore_rpg_controls=options.ignore_rpg_controls,
            )
            ok = needle in hay if options.mode == "keyword" else needle == hay
            if ok:
                matched.append(label)
        if matched:
            out.append(SearchResult(r, "＋".join(matched)))
    return out


def display_text(text: str, multiline: bool = False, limit: int | None = None) -> str:
    value = normalize_newlines(text or "")
    if not multiline:
        value = value.replace("\n", " ↵ ")
    if limit is not None and len(value) > limit:
        value = value[: max(0, limit - 1)] + "…"
    return value


def transform_outside_controls(text: str, mode: str) -> str:
    return _outside_controls_transform(text, lambda s: transform_alnum_segment(s, mode))


def transform_range_outside_controls(text: str, start: int, end: int, mode: str) -> str:
    start = max(0, min(len(text), start)); end = max(start, min(len(text), end))
    return text[:start] + transform_outside_controls(text[start:end], mode) + text[end:]


def split_preserving_controls(text: str) -> list[tuple[bool, str]]:
    out: list[tuple[bool, str]] = []
    last = 0
    for m in CONTROL_CODE_RE.finditer(text or ""):
        if m.start() > last:
            out.append((False, text[last:m.start()]))
        out.append((True, m.group(0)))
        last = m.end()
    if last < len(text or ""):
        out.append((False, text[last:]))
    return out


def wrap_text_by_units(text: str, limit: float) -> str:
    """Preview-only wrapping. RPG control codes have zero width and are preserved."""
    if limit <= 0:
        return text
    lines: list[str] = []
    for raw_line in normalize_newlines(text).split("\n"):
        current: list[str] = []
        units = 0.0
        for is_control, segment in split_preserving_controls(raw_line):
            if is_control:
                current.append(segment)
                continue
            for ch in segment:
                w = fullwidth_units(ch)
                if current and units + w > limit:
                    lines.append("".join(current)); current = []; units = 0.0
                current.append(ch); units += w
        lines.append("".join(current))
    return "\n".join(lines)


def analyze_speaker_groups(records: Sequence[DataRecord], pattern_mode: int = 0,
                           first_n: int | None = None) -> list[SpeakerGroup]:
    selected = [r for r in records if r.marker == "Message"]
    if first_n is not None:
        selected = selected[: max(0, first_n)]
    grouped: dict[str, dict[str, object]] = {}
    for r in selected:
        found = detect_speaker(r.original, pattern_mode)
        if not found:
            continue
        name, ptype, confidence = found
        tfound = detect_speaker(r.translated, ptype)
        tname = tfound[0] if tfound else ""
        g = grouped.setdefault(name, {
            "records": [], "types": set(), "confidences": [], "translations": defaultdict(lambda: {"count": 0, "types": set(), "confidences": []})
        })
        g["records"].append(r); g["types"].add(ptype); g["confidences"].append(confidence)
        tg = g["translations"][tname]
        tg["count"] += 1; tg["types"].add(ptype); tg["confidences"].append(confidence)
    result: list[SpeakerGroup] = []
    for original_name, g in grouped.items():
        options: list[SpeakerTranslationOption] = []
        for trans, tg in g["translations"].items():
            best = max(tg["confidences"], key=lambda x: CONFIDENCE_RANK.get(x, -1))
            options.append(SpeakerTranslationOption(trans, int(tg["count"]), tuple(sorted(tg["types"])), best))
        options.sort(key=lambda x: (-x.count, x.translation))
        max_conf = max(g["confidences"], key=lambda x: CONFIDENCE_RANK.get(x, -1))
        result.append(SpeakerGroup(
            original_name,
            tuple(options),
            len(g["records"]),
            tuple(sorted(g["types"])),
            max_conf,
            tuple(g["records"]),
        ))
    return sorted(result, key=lambda x: (-x.occurrences, x.original_name))


def quote_styles(text: str) -> tuple[str, ...]:
    found: list[str] = []
    tokens = {ch for _p, ch in _quote_tokens(text)}
    for label, (op, cl) in QUOTE_PAIRS.items():
        if op in tokens or cl in tokens:
            found.append(label)
    return tuple(found)


def _raw_line_starts(text: str) -> list[int]:
    starts = [0]
    starts.extend(i + 1 for i, ch in enumerate(text) if ch == "\n")
    return starts


def _insert_at_visible_line_start(text: str, line_index: int, char: str) -> str:
    lines = normalize_newlines(text).split("\n")
    if not lines:
        return text
    idx = max(0, min(line_index, len(lines) - 1))
    line = lines[idx]
    mask = _mask_controls(line)
    pos = next((i for i, ch in enumerate(line) if not mask[i] and not ch.isspace()), len(line))
    lines[idx] = line[:pos] + char + line[pos:]
    return "\n".join(lines)


def _replace_quote_char_at(text: str, pos: int, char: str) -> str:
    return text[:pos] + char + text[pos + 1:]


def _outer_quote_info(text: str) -> dict[str, object]:
    tokens = _quote_tokens(text); visible = _visible_positions(text); starts = _line_start_positions(text)
    info: dict[str, object] = {"tokens": tokens, "visible": visible, "styles": quote_styles(text)}
    if not visible:
        return info
    last = visible[-1]
    paired: list[tuple[int, int, str, str]] = []
    for op, cl in OPEN_TO_CLOSE.items():
        ops = [p for p, c in tokens if c == op]
        cls = [p for p, c in tokens if c == cl]
        if op == cl:
            for a, b in zip(ops[0::2], ops[1::2]): paired.append((a, b, op, cl))
        else:
            stack: list[int] = []
            for p, c in tokens:
                if c == op: stack.append(p)
                elif c == cl and stack: paired.append((stack.pop(0), p, op, cl))
    outer = next((x for x in paired if x[1] == last), None)
    if outer:
        a, b, op, cl = outer
        info.update({"open": op, "close": cl, "open_pos": a, "close_pos": b, "open_at_line_start": a in starts,
                     "open_line": normalize_newlines(text[:a]).count("\n"), "close_at_end": True, "paired": True})
        return info
    for p, ch in tokens:
        if ch in OPEN_TO_CLOSE and p in starts:
            cl = OPEN_TO_CLOSE[ch]
            later_close = any(c == cl and q > p for q, c in tokens)
            if not later_close:
                info.update({"open": ch, "close": None, "open_pos": p, "open_at_line_start": True,
                             "open_line": normalize_newlines(text[:p]).count("\n"), "paired": False})
                return info
    return info


def analyze_quote_rows(records: Iterable[DataRecord]) -> list[QuoteAnalysisRow]:
    rows: list[QuoteAnalysisRow] = []
    for r in records:
        if r.marker != "Message":
            continue
        sstyles = quote_styles(r.original); tstyles = quote_styles(r.translated)
        if not sstyles and not tstyles:
            continue
        sinfo = _outer_quote_info(r.original); tinfo = _outer_quote_info(r.translated)
        situation_eligible = bool(sinfo.get("open")) and (bool(sinfo.get("close_at_end")) or sinfo.get("close") is None)
        if situation_eligible:
            reason = "外层引号位置明确，可检查封闭状态；已有左引号位置将保持不变。"
            rows.append(QuoteAnalysisRow(r, "引号状况", "可批量", True, reason, sstyles, tstyles))
        else:
            rows.append(QuoteAnalysisRow(r, "引号状况", "手动", False, "外层引号不在安全位置，不能批量改变封闭状态。", sstyles, tstyles))
        if len(sstyles) == 1:
            rows.append(QuoteAnalysisRow(r, "引号样式", "可批量", True, "原文只使用一种引号，可统一译文引号样式。", sstyles, tstyles))
        else:
            rows.append(QuoteAnalysisRow(r, "引号样式", "手动", False, "原文没有引号或包含多种引号，不能自动决定统一样式。", sstyles, tstyles))
    return rows


def _replace_all_quote_styles(text: str, pair: tuple[str, str]) -> str:
    op_target, cl_target = pair
    chars = list(text); mask = _mask_controls(text)
    symmetric_state: dict[str, bool] = defaultdict(lambda: True)
    open_only = set(OPEN_TO_CLOSE) - set(OPEN_TO_CLOSE.values())
    close_only = set(OPEN_TO_CLOSE.values()) - set(OPEN_TO_CLOSE)
    symmetric = set(OPEN_TO_CLOSE) & set(OPEN_TO_CLOSE.values())
    for i, ch in enumerate(chars):
        if mask[i]:
            continue
        if ch in open_only:
            chars[i] = op_target
        elif ch in close_only:
            chars[i] = cl_target
        elif ch in symmetric:
            chars[i] = op_target if symmetric_state[ch] else cl_target
            symmetric_state[ch] = not symmetric_state[ch]
    return "".join(chars)


def build_quote_style_proposal(record: DataRecord, pair: tuple[str, str]) -> str:
    return _replace_all_quote_styles(record.translated, pair)


def _remove_outer_quote_positions(text: str) -> str:
    info = _outer_quote_info(text)
    positions = []
    if isinstance(info.get("open_pos"), int): positions.append(int(info["open_pos"]))
    if isinstance(info.get("close_pos"), int): positions.append(int(info["close_pos"]))
    chars = list(text)
    for pos in sorted(set(positions), reverse=True):
        if 0 <= pos < len(chars): del chars[pos]
    return "".join(chars)


def build_quote_situation_proposal(record: DataRecord, form: str,
                                   pair: tuple[str, str] | None = None) -> str | None:
    """Change only the outer quote condition. Existing opening quote position is never moved."""
    source = _outer_quote_info(record.original)
    if not source.get("open"):
        return None
    target_pair = pair or (str(source["open"]), str(source.get("close") or OPEN_TO_CLOSE[str(source["open"])]))
    op, cl = target_pair
    text = record.translated
    target = _outer_quote_info(text)
    if form == "删除引号":
        return _remove_outer_quote_positions(text)

    # Opening quote: replace it in place if present; otherwise insert only at the source line start.
    open_pos = target.get("open_pos")
    if not isinstance(open_pos, int):
        # Preserve any existing opening-quote position even when its pair/style is malformed.
        open_pos = next((p for p, ch in _quote_tokens(text) if ch in OPEN_TO_CLOSE), None)
    if isinstance(open_pos, int):
        text = _replace_quote_char_at(text, open_pos, op)
    else:
        if not source.get("open_at_line_start"):
            return None
        text = _insert_at_visible_line_start(text, int(source.get("open_line", 0)), op)

    # Recompute after possible insertion.
    target = _outer_quote_info(text)
    if form == "只有左引号":
        close_pos = target.get("close_pos")
        if not isinstance(close_pos, int):
            vis = _visible_positions(text)
            close_pos = vis[-1] if vis and text[vis[-1]] in ALL_QUOTES else None
        if isinstance(close_pos, int):
            text = text[:close_pos] + text[close_pos + 1:]
        return text

    if form == "左右引号":
        close_pos = target.get("close_pos")
        if not isinstance(close_pos, int):
            vis = _visible_positions(text)
            # A wrong-style closing quote at the visible end is replaced in place, never duplicated.
            close_pos = vis[-1] if vis and text[vis[-1]] in ALL_QUOTES else None
        if isinstance(close_pos, int):
            return _replace_quote_char_at(text, close_pos, cl)
        vis = _visible_positions(text)
        if not vis:
            return text
        end = vis[-1] + 1
        return text[:end] + cl + text[end:]
    return text


def default_symbol_catalog() -> list[SymbolCatalogEntry]:
    groups = [
        ("点号", [("。", "句号", [".", "．"]), ("？", "问号", ["?"]), ("！", "叹号", ["!"]),
                  ("，", "逗号", [",", "、"]), ("、", "顿号", ["，", ","]), ("；", "分号", [";"]), ("：", "冒号", [":"])]),
        ("标号", [("“", "左双引号", ["「", "『", '"']), ("”", "右双引号", ["」", "』", '"']),
                  ("‘", "左单引号", ["'", "「"]), ("’", "右单引号", ["'", "」"]),
                  ("（", "左圆括号", ["("]), ("）", "右圆括号", [")"]), ("[", "左方括号", ["［", "【"]), ("]", "右方括号", ["］", "】"]),
                  ("{", "左花括号", ["｛"]), ("}", "右花括号", ["｝"]), ("——", "破折号", ["—"]),
                  ("……", "省略号", ["…", "..."]), ("．", "着重号", ["."]), ("《", "左书名号", ["〈"]), ("》", "右书名号", ["〉"]),
                  ("·", "间隔号", ["・"]), ("—", "连接号", ["－", "-"]), ("____", "专名号", ["＿"]), ("/", "分隔号", ["／"])])
    ]
    result: list[SymbolCatalogEntry] = []
    # fix a typographic quote in source literal defensively
    for category, items in groups:
        for symbol, desc, peers in items:
            result.append(SymbolCatalogEntry(str(symbol), category, desc, tuple(str(x) for x in peers)))
    return result


def save_symbol_catalog(path: Path, entries: Sequence[SymbolCatalogEntry]) -> None:
    payload = [{"symbol": e.symbol, "category": e.category, "description": e.description, "peers": list(e.peers)} for e in entries]
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def load_symbol_catalog(path: Path) -> list[SymbolCatalogEntry]:
    if not path.exists():
        entries = default_symbol_catalog(); save_symbol_catalog(path, entries); return entries
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
        return [SymbolCatalogEntry(str(x["symbol"]), str(x.get("category", "其他")), str(x.get("description", "")), tuple(map(str, x.get("peers", [])))) for x in payload]
    except Exception as exc:
        raise QAError(f"标点分类文档无法读取：{path}\n{exc}") from exc


def punctuation_text(text: str) -> str:
    # Control operators are excluded. Literal punctuation outside operators remains countable.
    return strip_control_codes(normalize_newlines(text))


def analyze_symbol_catalog(records: Iterable[DataRecord], catalog: Sequence[SymbolCatalogEntry]) -> list[dict[str, object]]:
    oc: Counter[str] = Counter(); tc: Counter[str] = Counter()
    originals: list[str] = []; translations: list[str] = []
    for r in records:
        otext = punctuation_text(r.original); ttext = punctuation_text(r.translated)
        originals.append(otext); translations.append(ttext)
        for ch in otext:
            if unicodedata.category(ch).startswith("P") or ch in ALL_QUOTES: oc[ch] += 1
        for ch in ttext:
            if unicodedata.category(ch).startswith("P") or ch in ALL_QUOTES: tc[ch] += 1
    # Catalogue entries may intentionally be multi-character units such as —— or …….
    for entry in catalog:
        if len(entry.symbol) > 1:
            oc[entry.symbol] = sum(text.count(entry.symbol) for text in originals)
            tc[entry.symbol] = sum(text.count(entry.symbol) for text in translations)
    mapping = {e.symbol: e for e in catalog}
    discovered = sorted(set(oc) | set(tc) | set(mapping), key=lambda x: (mapping.get(x, SymbolCatalogEntry(x, "其他", "未分类", ())).category, x))
    rows: list[dict[str, object]] = []
    categories: dict[str, list[str]] = defaultdict(list)
    for e in catalog: categories[e.category].append(e.symbol)
    for symbol in discovered:
        e = mapping.get(symbol, SymbolCatalogEntry(symbol, "其他", "未分类", ()))
        same_class = [x for x in e.peers if x != symbol]
        same_group = [x for x in categories.get(e.category, []) if x != symbol and x not in same_class]
        choices = [symbol] + same_class + ["自定义", "删除"] + same_group
        rows.append({
            "symbol": symbol, "category": e.category, "description": e.description,
            "unicode": " ".join(f"U+{ord(ch):04X}" for ch in symbol), "choices": choices,
            "original_count": oc[symbol], "translated_count": tc[symbol], "total_count": oc[symbol] + tc[symbol],
        })
    return rows


def analyze_alnum_occurrences(records: Iterable[DataRecord], target: str = "both") -> list[TextChange]:
    if target == "english": pattern = re.compile(r"[A-Za-zＡ-Ｚａ-ｚ]+")
    elif target == "digits": pattern = re.compile(r"[0-9０-９]+")
    else: pattern = re.compile(r"[A-Za-zＡ-Ｚａ-ｚ0-9０-９]+")
    out: list[TextChange] = []
    for r in records:
        values = []
        for is_control, seg in split_preserving_controls(r.translated):
            if not is_control: values.extend(m.group(0) for m in pattern.finditer(seg))
        if values:
            out.append(TextChange(r, "、".join(values[:20]), r.translated))
    return out


def _ellipsis_visible_stream(text: str) -> list[tuple[str, int, int, bool]]:
    """Return visible point/ellipsis tokens and whether control codes occurred since prior token."""
    tokens: list[tuple[str, int, int, bool]] = []
    last = 0; control_since = False
    for m in CONTROL_CODE_RE.finditer(text):
        segment = text[last:m.start()]
        for j, ch in enumerate(segment, last):
            if ch in ".．·・。…": tokens.append((ch, j, j + 1, control_since)); control_since = False
            elif not ch.isspace(): control_since = False
        control_since = True
        last = m.end()
    for j, ch in enumerate(text[last:], last):
        if ch in ".．·・。…": tokens.append((ch, j, j + 1, control_since)); control_since = False
        elif not ch.isspace(): control_since = False
    return tokens


def ellipsis_occurrences_for_text(record: DataRecord, side: str, text: str) -> list[EllipsisOccurrence]:
    tokens = _ellipsis_visible_stream(text)
    if not tokens: return []
    groups: list[list[tuple[str,int,int,bool]]] = []
    current: list[tuple[str,int,int,bool]] = []
    prev_end = -1
    for tok in tokens:
        ch, start, end, interrupted = tok
        between = text[prev_end:start] if prev_end >= 0 else ""
        continuation = bool(current) and (not strip_control_codes(between).strip())
        if continuation: current.append(tok)
        else:
            if current: groups.append(current)
            current = [tok]
        prev_end = end
    if current: groups.append(current)
    out: list[EllipsisOccurrence] = []
    styles_in_text = set()
    for group in groups:
        chars = "".join(x[0] for x in group); interrupted = any(x[3] for x in group[1:]) or bool(CONTROL_CODE_RE.search(text[group[0][1]:group[-1][2]]))
        if all(ch == "…" for ch in chars):
            kind = "省略号"; point_char = "…"; count = len(chars); style = "…" * len(chars)
        else:
            kind = "连续点"; point_char = chars[0] if len(set(chars)) == 1 else "混合点"; count = len(chars); style = chars
        styles_in_text.add((kind, point_char))
        out.append(EllipsisOccurrence(record, side, text[group[0][1]:group[-1][2]], style, kind, point_char, count,
                                      interrupted, group[0][1], group[-1][2]))
    if side == "原文" and len(styles_in_text) > 1:
        out = [EllipsisOccurrence(x.record, x.side, x.raw, x.visible_style, x.kind, x.point_char, x.count,
                                  x.interrupted_by_control, x.start, x.end, True, "原文包含多种省略号/连续点形式") for x in out]
    return out


def analyze_ellipsis_occurrences(records: Iterable[DataRecord], side: str = "both") -> list[EllipsisOccurrence]:
    out: list[EllipsisOccurrence] = []
    for r in records:
        source_rows = ellipsis_occurrences_for_text(r, "原文", r.original)
        source_styles = {(x.kind, x.point_char) for x in source_rows}
        source_manual = len(source_styles) > 1
        translated_rows = ellipsis_occurrences_for_text(r, "译文", r.translated)
        translated_styles = {(x.kind, x.point_char) for x in translated_rows}
        style_mismatch = bool(source_styles and translated_styles and source_styles != translated_styles)
        if source_manual:
            source_rows = [EllipsisOccurrence(x.record, x.side, x.raw, x.visible_style, x.kind, x.point_char, x.count,
                                              x.interrupted_by_control, x.start, x.end, True, "原文包含多种省略号/连续点形式") for x in source_rows]
            translated_rows = [EllipsisOccurrence(x.record, x.side, x.raw, x.visible_style, x.kind, x.point_char, x.count,
                                                  x.interrupted_by_control, x.start, x.end, True, "对应原文包含多种省略号形式，只能手动处理") for x in translated_rows]
        elif style_mismatch:
            translated_rows = [EllipsisOccurrence(x.record, x.side, x.raw, x.visible_style, x.kind, x.point_char, x.count,
                                                  x.interrupted_by_control, x.start, x.end, x.manual_only, "原文与译文省略号形式不一致") for x in translated_rows]
        if side in {"original", "both"}: out.extend(source_rows)
        if side in {"translated", "both"}: out.extend(translated_rows)
    return out


def _replace_occurrences_reverse(text: str, replacements: Sequence[tuple[int, int, str]]) -> str:
    for start, end, repl in sorted(replacements, reverse=True): text = text[:start] + repl + text[end:]
    return text


def build_ellipsis_conversion(record: DataRecord, *, direction: str, source_chars: set[str] | None = None,
                              group_size: int = 2, remainder: str = "删除", ellipsis_style: str = "……",
                              dot_style: str = ".", max_ellipsis: int = 2) -> str | None:
    occs = ellipsis_occurrences_for_text(record, "译文", record.translated)
    replacements: list[tuple[int,int,str]] = []
    for x in occs:
        if direction == "连续点→省略号" and x.kind == "连续点":
            if source_chars and x.point_char not in source_chars and "混合点" not in source_chars: continue
            if x.interrupted_by_control:
                if group_size != 1: continue
                raw = record.translated[x.start:x.end]
                repl = _outside_controls_transform(raw, lambda seg: "".join(ellipsis_style if ch in ".．·・。" else ch for ch in seg))
                replacements.append((x.start, x.end, repl)); continue
            full, rem = divmod(x.count, max(1, group_size))
            repl = ellipsis_style * full
            if rem and remainder == "一个省略号": repl += ellipsis_style
            replacements.append((x.start, x.end, repl))
        elif direction == "省略号→连续点" and x.kind == "省略号":
            if x.interrupted_by_control:
                raw = record.translated[x.start:x.end]
                repl = _outside_controls_transform(raw, lambda seg: seg.replace("…", dot_style))
                replacements.append((x.start, x.end, repl)); continue
            replacements.append((x.start, x.end, dot_style * x.count))
        elif direction == "省略号压缩" and x.kind == "省略号" and x.count > max_ellipsis:
            replacements.append((x.start, x.end, "…" * max_ellipsis))
        elif direction == "单省略号→双省略号" and x.kind == "省略号" and x.count == 1:
            replacements.append((x.start, x.end, "……"))
    if not replacements: return None
    return _replace_occurrences_reverse(record.translated, replacements)


def analyze_duplicate_texts(records: Iterable[DataRecord]) -> list[DuplicateTextGroup]:
    groups: dict[str, list[DataRecord]] = defaultdict(list)
    for r in records:
        groups[r.original].append(r)
    result: list[DuplicateTextGroup] = []
    for original, recs in groups.items():
        if len(recs) < 2: continue
        counts = Counter(r.translated for r in recs)
        result.append(DuplicateTextGroup(original, tuple(sorted(counts.items(), key=lambda x: (-x[1], x[0]))), tuple(recs)))
    return sorted(result, key=lambda x: (-len(x.records), x.original))


def duplicate_groups_to_dictionary(groups: Sequence[DuplicateTextGroup], only_consistent: bool = False) -> list[EditableDictionaryRow]:
    rows: list[EditableDictionaryRow] = []
    for group in groups:
        if only_consistent and not group.consistent: continue
        for translation, _count in group.translations:
            rows.append(EditableDictionaryRow(group.original, translation, "其他", "重复文本"))
    return rows

# ============================================================================
# v4.1 additions: RM2k/2k3 control-code awareness, combined quote analysis,
# refined ellipsis rules and source-format profiles.
# ============================================================================
from dataclasses import replace as dataclass_replace

# RPG Maker 2000/2003 message codes. We intentionally do not add MV/MZ-only
# controls such as \I[n], \{, \}, \PX[n], \PY[n], or \FS[n].
RM2K3_CONTROL_RE = re.compile(
    r"(?:\\(?:[CSNVcsnv]\[(?:\\[Vv]\[\d+\]|[^\]\r\n]+)\]|[\$!\.\|\^><\\_])|\$[A-Za-z])"
)

RM2K3_SYMBOL_GLYPHS = {
    "$A": "⚔", "$B": "🛡", "$C": "✡", "$D": "☀", "$E": "☾",
    "$F": "☿", "$G": "♀", "$H": "♀", "$I": "♂", "$J": "♃",
    "$K": "♄", "$L": "♅", "$M": "♆", "$N": "♇", "$O": "♈",
    "$P": "♉", "$Q": "♊", "$R": "♋", "$S": "♌", "$T": "♍",
    "$U": "♎", "$V": "♏", "$W": "♐", "$X": "♑", "$Y": "♒",
    "$Z": "♓", "$a": "☺", "$b": "😐", "$c": "☹", "$d": "💧",
    "$e": "💦", "$f": "♤", "$g": "♡", "$h": "♢", "$i": "♧",
    "$j": "♠", "$k": "♥", "$l": "♦", "$m": "♣", "$n": "☠",
    "$o": "×", "$p": "☀", "$q": "☾", "$r": "·", "$s": "↑",
    "$t": "→", "$u": "↓", "$v": "←", "$w": "↗", "$x": "↘",
    "$y": "↙", "$z": "↖",
}


def iter_rm2k3_controls(text: str) -> list[tuple[int, int, str]]:
    return [(m.start(), m.end(), m.group(0)) for m in RM2K3_CONTROL_RE.finditer(text or "")]


def strip_control_codes(text: str) -> str:
    """Remove RM2k/2k3 control operators and legacy-recognized controls.

    This supersedes the imported legacy helper so v4.1 also protects ``\\_`` and
    the $A-$z special-character codes.
    """
    value = text or ""
    value = RM2K3_CONTROL_RE.sub("", value)
    return CONTROL_CODE_RE.sub("", value)


def render_rm2k3_controls(text: str, *, runtime_placeholders: bool = True) -> str:
    """Hide operators while retaining what can be shown statically.

    Static special-character codes ($A-$z) are rendered with readable Unicode
    approximations. Runtime-dependent hero/variable substitutions are shown as
    placeholders when requested; pure timing/style controls disappear. Literal
    backslashes produced by ``\\`` are protected from the legacy matcher, so
    hiding operators never eats the ordinary text that follows them.
    """
    out: list[str] = []
    for is_control, segment in split_preserving_controls(text or ""):
        if not is_control:
            out.append(segment)
            continue
        code = segment
        low = code.lower()
        repl = ""
        if code in RM2K3_SYMBOL_GLYPHS:
            repl = RM2K3_SYMBOL_GLYPHS[code]
        elif code == r"\\":
            repl = "\\"
        elif low == r"\_":
            repl = "\u2009"  # approximate half-space visually
        elif low.startswith(r"\n[") and runtime_placeholders:
            repl = "〔角色名" + code[3:-1] + "〕"
        elif low.startswith(r"\v[") and runtime_placeholders:
            repl = "〔变量" + code[3:-1] + "〕"
        # \C, \S and timing/input controls intentionally produce no glyph.
        out.append(repl)
    return "".join(out)


def split_preserving_controls(text: str) -> list[tuple[bool, str]]:
    out: list[tuple[bool, str]] = []
    value = text or ""
    last = 0
    for m in RM2K3_CONTROL_RE.finditer(value):
        if m.start() > last:
            # Legacy regex might still find another control in this segment.
            segment = value[last:m.start()]
            sublast = 0
            for lm in CONTROL_CODE_RE.finditer(segment):
                if lm.start() > sublast:
                    out.append((False, segment[sublast:lm.start()]))
                out.append((True, lm.group(0)))
                sublast = lm.end()
            if sublast < len(segment):
                out.append((False, segment[sublast:]))
        out.append((True, m.group(0)))
        last = m.end()
    if last < len(value):
        segment = value[last:]
        sublast = 0
        for lm in CONTROL_CODE_RE.finditer(segment):
            if lm.start() > sublast:
                out.append((False, segment[sublast:lm.start()]))
            out.append((True, lm.group(0)))
            sublast = lm.end()
        if sublast < len(segment):
            out.append((False, segment[sublast:]))
    return out


def _outside_controls_transform(text: str, fn: Callable[[str], str]) -> str:
    return "".join(segment if is_control else fn(segment) for is_control, segment in split_preserving_controls(text or ""))


def punctuation_text(text: str) -> str:
    # Render RM2k/2k3 external glyphs first, so e.g. $r contributes a visible
    # middle-dot to punctuation statistics, while the operator itself is absent.
    return render_rm2k3_controls(normalize_newlines(text or ""), runtime_placeholders=False)


@dataclass(frozen=True)
class CombinedQuoteAnalysisRow:
    record: DataRecord
    situation: str
    style: str
    situation_match: str
    style_match: str
    situation_auto: bool
    style_auto: bool
    situation_reason: str
    style_reason: str
    source_styles: tuple[str, ...]
    translated_styles: tuple[str, ...]
    proposal: str | None = None


def _quote_situation_label(info: Mapping[str, object], styles: tuple[str, ...]) -> str:
    if info.get("open") and info.get("close_at_end"):
        where = "行首左引号" if info.get("open_at_line_start") else "行中左引号"
        return f"左右封闭（{where}，右引号在末尾）"
    if info.get("open") and not info.get("close"):
        where = "行首" if info.get("open_at_line_start") else "行中"
        return f"只有左引号（{where}）"
    if styles:
        return "存在内部/复杂引号（手动）"
    return "无引号"


def _quote_situation_signature(info: Mapping[str, object]) -> tuple[bool, bool, bool]:
    return (bool(info.get("open")), bool(info.get("close_at_end")), bool(info.get("open_at_line_start")))


def analyze_quote_combined(records: Iterable[DataRecord]) -> list[CombinedQuoteAnalysisRow]:
    rows: list[CombinedQuoteAnalysisRow] = []
    for r in records:
        if r.marker != "Message":
            continue
        ss = quote_styles(r.original); ts = quote_styles(r.translated)
        if not ss and not ts:
            continue
        si = _outer_quote_info(r.original); ti = _outer_quote_info(r.translated)
        s_label = _quote_situation_label(si, ss); t_label = _quote_situation_label(ti, ts)
        situation_match = "是" if _quote_situation_signature(si) == _quote_situation_signature(ti) else "否"
        style_match = "是" if ss == ts else "否"
        situation_auto = bool(si.get("open")) and (bool(si.get("close_at_end")) or si.get("close") is None)
        style_auto = len(ss) == 1
        s_reason = ("原文外层引号位置明确，可批量处理；已有译文左引号位置保持不动。"
                    if situation_auto else "原文外层引号位置不满足安全批处理条件，需要手动处理。")
        st_reason = ("原文只包含一种引号样式，可批量统一译文样式。"
                     if style_auto else "原文包含零种或多种引号样式，不能自动决定样式。")
        rows.append(CombinedQuoteAnalysisRow(
            r,
            f"原文：{s_label}；译文：{t_label}",
            f"原文：{'、'.join(ss) or '无'}；译文：{'、'.join(ts) or '无'}",
            situation_match,
            style_match,
            situation_auto,
            style_auto,
            s_reason,
            st_reason,
            ss,
            ts,
        ))
    return rows


def case_profile(text: str) -> str:
    visible = render_rm2k3_controls(text or "", runtime_placeholders=False)
    words = re.findall(r"[A-Za-z]+", unicodedata.normalize("NFKC", visible))
    if not words:
        return "无英文"
    if all(w.isupper() for w in words):
        return "原文大写"
    if all(w.islower() for w in words):
        return "原文小写"
    if all((len(w) == 1 and w.isupper()) or (w[:1].isupper() and w[1:].islower()) for w in words):
        return "原文首字母大写"
    return "原文大小写混搭"


def case_format_matches(original: str, translated: str) -> bool:
    op = case_profile(original)
    if op == "无英文":
        return True
    tp = case_profile(translated).replace("原文", "原文")
    return tp == op


def width_profile(text: str, target: str = "both") -> str:
    value = render_rm2k3_controls(text or "", runtime_placeholders=False)
    chars: list[str] = []
    if target in {"english", "both"}:
        chars.extend(re.findall(r"[A-Za-zＡ-Ｚａ-ｚ]", value))
    if target in {"digits", "both"}:
        chars.extend(re.findall(r"[0-9０-９]", value))
    if not chars:
        return "无目标字符"
    half = any(("A" <= ch <= "Z") or ("a" <= ch <= "z") or ("0" <= ch <= "9") for ch in chars)
    full = any(("Ａ" <= ch <= "Ｚ") or ("ａ" <= ch <= "ｚ") or ("０" <= ch <= "９") for ch in chars)
    if half and full:
        return "原文混搭"
    return "原文全角" if full else "原文半角"


def width_format_matches(original: str, translated: str, target: str = "both") -> bool:
    op = width_profile(original, target)
    if op == "无目标字符":
        return True
    return width_profile(translated, target) == op


# Refined ellipsis analysis: a continuous-point run means at least two of the
# SAME point character. The ellipsis character (…) is its own category and is
# never classified as a continuous point.
POINT_CHARS = ".．·・。"


def _ellipsis_tokens(text: str) -> list[tuple[str, int, int, bool]]:
    value = text or ""
    controls = [(s, e) for s, e, _ in iter_rm2k3_controls(value)]
    # Also include broader legacy controls not already covered.
    controls += [(m.start(), m.end()) for m in CONTROL_CODE_RE.finditer(value)]
    controls = sorted(set(controls))
    hidden = [False] * len(value)
    for s, e in controls:
        for i in range(max(0, s), min(len(value), e)):
            hidden[i] = True
    tokens: list[tuple[str, int, int, bool]] = []
    control_since = False
    i = 0
    while i < len(value):
        if hidden[i]:
            control_since = True; i += 1; continue
        ch = value[i]
        if ch in POINT_CHARS or ch == "…":
            tokens.append((ch, i, i + 1, control_since)); control_since = False
        elif not ch.isspace():
            control_since = False
        i += 1
    return tokens


def ellipsis_occurrences_for_text(record: DataRecord, side: str, text: str) -> list[EllipsisOccurrence]:
    tokens = _ellipsis_tokens(text)
    if not tokens:
        return []
    out: list[EllipsisOccurrence] = []
    i = 0
    while i < len(tokens):
        ch, start, end, interrupted = tokens[i]
        if ch == "…":
            j = i + 1
            any_control = False
            while j < len(tokens) and tokens[j][0] == "…":
                between = text[tokens[j-1][2]:tokens[j][1]]
                if strip_control_codes(between).strip():
                    break
                any_control = any_control or tokens[j][3] or bool(RM2K3_CONTROL_RE.search(between)) or bool(CONTROL_CODE_RE.search(between))
                j += 1
            count = j - i
            raw = text[start:tokens[j-1][2]]
            out.append(EllipsisOccurrence(record, side, raw, "…" * count, "省略号", "…", count,
                                          interrupted or any_control, start, tokens[j-1][2]))
            i = j
            continue
        # Continuous points require the same point character and at least 2.
        j = i + 1; any_control = False
        while j < len(tokens) and tokens[j][0] == ch:
            between = text[tokens[j-1][2]:tokens[j][1]]
            if strip_control_codes(between).strip():
                break
            any_control = any_control or tokens[j][3] or bool(RM2K3_CONTROL_RE.search(between)) or bool(CONTROL_CODE_RE.search(between))
            j += 1
        count = j - i
        if count >= 2:
            raw = text[start:tokens[j-1][2]]
            out.append(EllipsisOccurrence(record, side, raw, ch * count, "连续点", ch, count,
                                          interrupted or any_control, start, tokens[j-1][2]))
        i = max(j, i + 1)
    return out


def _ellipsis_summary(rows: Sequence[EllipsisOccurrence]) -> tuple[tuple[str, str], ...]:
    return tuple((x.kind, x.point_char) for x in rows)


def _ellipsis_count_summary(rows: Sequence[EllipsisOccurrence]) -> tuple[tuple[str, str, int], ...]:
    return tuple((x.kind, x.point_char, x.count) for x in rows)


def analyze_ellipsis_occurrences(records: Iterable[DataRecord], side: str = "both") -> list[EllipsisOccurrence]:
    out: list[EllipsisOccurrence] = []
    for r in records:
        source_rows = ellipsis_occurrences_for_text(r, "原文", r.original)
        trans_rows = ellipsis_occurrences_for_text(r, "译文", r.translated)
        source_forms = {(x.kind, x.point_char) for x in source_rows}
        source_manual = len(source_forms) > 1
        form_mismatch = bool(source_rows or trans_rows) and _ellipsis_summary(source_rows) != _ellipsis_summary(trans_rows)
        count_mismatch = (not form_mismatch and bool(source_rows or trans_rows)
                          and _ellipsis_count_summary(source_rows) != _ellipsis_count_summary(trans_rows))
        if source_manual:
            source_rows = [dataclass_replace(x, manual_only=True, reason="原文存在多种省略号/连续点形式，只能手动处理") for x in source_rows]
            trans_rows = [dataclass_replace(x, manual_only=True, reason="对应原文存在多种省略号/连续点形式，只能手动处理") for x in trans_rows]
        elif form_mismatch:
            trans_rows = [dataclass_replace(x, reason="原文与译文省略号形式不一致") for x in trans_rows]
        elif count_mismatch:
            trans_rows = [dataclass_replace(x, reason="原文与译文省略号数量不一致") for x in trans_rows]
        if side in {"original", "both"}: out.extend(source_rows)
        if side in {"translated", "both"}: out.extend(trans_rows)
    return out


def build_ellipsis_conversion(record: DataRecord, *, direction: str, source_chars: set[str] | None = None,
                              group_size: int = 2, remainder: str = "删除", ellipsis_style: str = "……",
                              dot_style: str = ".", dots_per_ellipsis: int = 3,
                              target_ellipsis_count: int = 2, max_ellipsis: int = 2) -> str | None:
    occs = ellipsis_occurrences_for_text(record, "译文", record.translated)
    replacements: list[tuple[int, int, str]] = []
    for x in occs:
        if direction == "连续点→省略号" and x.kind == "连续点":
            if source_chars and x.point_char not in source_chars:
                continue
            if x.interrupted_by_control:
                # Only 1 point -> 1 ellipsis is safe with embedded controls.
                if group_size != 1:
                    continue
                raw = record.translated[x.start:x.end]
                repl = _outside_controls_transform(raw, lambda seg: "".join(ellipsis_style if ch == x.point_char else ch for ch in seg))
                replacements.append((x.start, x.end, repl)); continue
            full, rem = divmod(x.count, max(1, group_size))
            repl = ellipsis_style * full
            if rem and remainder == "一个省略号":
                repl += ellipsis_style
            replacements.append((x.start, x.end, repl))
        elif direction == "省略号→连续点" and x.kind == "省略号":
            each = max(1, dots_per_ellipsis)
            if x.interrupted_by_control:
                raw = record.translated[x.start:x.end]
                repl = _outside_controls_transform(raw, lambda seg: seg.replace("…", dot_style * each))
            else:
                repl = dot_style * (x.count * each)
            replacements.append((x.start, x.end, repl))
        elif direction == "省略号→省略号" and x.kind == "省略号":
            target = max(1, target_ellipsis_count)
            replacements.append((x.start, x.end, "…" * target))
        elif direction == "省略号压缩" and x.kind == "省略号" and x.count > max_ellipsis:
            replacements.append((x.start, x.end, "…" * max_ellipsis))
        elif direction == "单省略号→双省略号" and x.kind == "省略号" and x.count == 1:
            replacements.append((x.start, x.end, "……"))
    if not replacements:
        return None
    return _replace_occurrences_reverse(record.translated, replacements)
