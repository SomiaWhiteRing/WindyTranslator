from __future__ import annotations

import csv
import argparse
import json
import math
import re
import shutil
import sys
import tkinter as tk
from dataclasses import replace as dataclass_replace
from pathlib import Path
from tkinter import filedialog, font as tkfont, messagebox, simpledialog, ttk
from typing import Iterable

from openpyxl import load_workbook

from workbench_app_v4 import RPGMakerProofreadingApp as V4App, DICT_CATEGORIES, QUOTE_STYLE_LABELS
from workbench_core import (
    CombinedQuoteAnalysisRow,
    DataRecord,
    EditableDictionaryRow,
    EllipsisOccurrence,
    QAError,
    QUOTE_PAIRS,
    TextChange,
    _outer_quote_info,
    analyze_alnum_occurrences,
    analyze_ellipsis_occurrences,
    analyze_quote_combined,
    build_ellipsis_conversion,
    build_quote_situation_proposal,
    build_quote_style_proposal,
    build_used_dictionary,
    case_format_matches,
    case_profile,
    dictionary_entries_from_rows,
    load_editable_dictionary,
    open_file,
    render_rm2k3_controls,
    save_dictionary,
    strip_control_codes,
    transform_outside_controls,
    width_format_matches,
    width_profile,
)

APP_TITLE = "RPG制作大师校对工具 v4.1"
CONFIG_NAME = "rpg_maker_proofreading_tool_config_v41.json"

TEXT_FILTER_COLUMNS = {
    "original", "translated", "proposal", "proposed", "reason", "file", "baseline",
    "term", "expected", "translation", "files", "found", "description", "custom",
    "replacement_desc", "symbol", "original_name",
}
FIXED_FILTER_COLUMNS = {
    "use", "status", "mode", "marker", "type", "side", "kind", "manual", "control",
    "field", "category", "confidence", "match", "situation_match", "style_match",
    "source_profile", "translated_profile", "face_type", "batch", "group",
}


class MultiChoiceFilterDialog(tk.Toplevel):
    def __init__(self, parent, title: str, values: list[str]):
        super().__init__(parent)
        self.title(title); self.transient(parent); self.grab_set(); self.geometry("430x560")
        self.result: set[str] | None = None
        self.vars: dict[str, tk.BooleanVar] = {v: tk.BooleanVar(value=True) for v in values}
        top = ttk.Frame(self, padding=8); top.pack(fill="both", expand=True)
        actions = ttk.Frame(top); actions.pack(fill="x")
        ttk.Button(actions, text="全选", command=lambda: self._set_all(True)).pack(side="left")
        ttk.Button(actions, text="全不选", command=lambda: self._set_all(False)).pack(side="left", padx=5)
        canvas = tk.Canvas(top, highlightthickness=0)
        scroll = ttk.Scrollbar(top, orient="vertical", command=canvas.yview)
        inner = ttk.Frame(canvas)
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=scroll.set)
        canvas.pack(side="left", fill="both", expand=True, pady=8); scroll.pack(side="right", fill="y", pady=8)
        for value in values:
            label = value if value else "（空白）"
            ttk.Checkbutton(inner, text=label, variable=self.vars[value]).pack(anchor="w", padx=4, pady=2)
        bottom = ttk.Frame(self, padding=(8, 0, 8, 8)); bottom.pack(fill="x")
        ttk.Button(bottom, text="确定", command=self._ok).pack(side="right")
        ttk.Button(bottom, text="取消", command=self.destroy).pack(side="right", padx=5)
        self.bind("<Escape>", lambda e: self.destroy())
        self.wait_visibility(); self.focus_set()

    def _set_all(self, state: bool):
        for var in self.vars.values(): var.set(state)

    def _ok(self):
        self.result = {v for v, var in self.vars.items() if var.get()}
        self.destroy()


class CategoryDialog(tk.Toplevel):
    def __init__(self, parent, filename: str):
        super().__init__(parent)
        self.title("设置辞典类别"); self.transient(parent); self.grab_set(); self.resizable(False, False)
        self.result: str | None = None
        ttk.Label(self, text=f"辞典没有“类别”列：\n{filename}\n请选择整个辞典使用的类别：", justify="left").pack(anchor="w", padx=12, pady=(12, 6))
        self.var = tk.StringVar(value="其他")
        combo = ttk.Combobox(self, textvariable=self.var, values=DICT_CATEGORIES, width=28)
        combo.pack(fill="x", padx=12); combo.focus_set()
        bar = ttk.Frame(self); bar.pack(fill="x", padx=12, pady=12)
        ttk.Button(bar, text="确定", command=self._ok).pack(side="right")
        ttk.Button(bar, text="取消", command=self.destroy).pack(side="right", padx=5)
        self.wait_visibility()

    def _ok(self):
        self.result = self.var.get().strip() or "其他"
        self.destroy()


class RPGMakerProofreadingApp(V4App):
    @staticmethod
    def _app_dir() -> Path:
        # Source mode: the script directory.  PyInstaller one-file mode: keep
        # editable configuration and reference documents beside the EXE rather
        # than inside the temporary _MEIPASS directory.
        if getattr(sys, "frozen", False):
            return Path(sys.executable).resolve().parent
        return Path(__file__).resolve().parent

    @staticmethod
    def _bundle_dir() -> Path:
        return Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))

    def _ensure_external_resource(self, relative: str):
        target = self._app_dir() / relative
        if target.exists():
            return
        source = self._bundle_dir() / relative
        if source.exists():
            target.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(source, target)

    def __init__(self, initial_input=None, initial_origin_dir=None, initial_translated_dir=None):
        super().__init__(
            initial_input=initial_input,
            initial_origin_dir=initial_origin_dir,
            initial_translated_dir=initial_translated_dir,
        )
        self.title(APP_TITLE)
        self.config_path = self._app_dir() / CONFIG_NAME
        self.config_dir = self._app_dir() / "配置"
        self.reference_dir = self._app_dir() / "资料"
        self.config_dir.mkdir(exist_ok=True); self.reference_dir.mkdir(exist_ok=True)
        # In packaged builds the defaults are bundled, then copied beside the
        # EXE on first run so the user can open and edit them normally.
        for rel in (
            "配置/标点符号分类.json",
            "资料/RPG_Maker_2000_2003_操作符一览.json",
            "资料/RPG_Maker_2000_2003_操作符一览.txt",
        ):
            self._ensure_external_resource(rel)
        self.catalog_path = self.config_dir / "标点符号分类.json"
        # Compatibility migration from v4.0's root-level catalogue.
        old_catalog = self._app_dir() / "标点符号分类.json"
        if not self.catalog_path.exists() and old_catalog.exists():
            self.catalog_path.write_bytes(old_catalog.read_bytes())
        self.control_catalog_path = self.reference_dir / "RPG_Maker_2000_2003_操作符一览.json"
        self.control_doc_path = self.reference_dir / "RPG_Maker_2000_2003_操作符一览.txt"
        # Parent v4 loads its own config filename during super().__init__().
        # Reload after switching to the v4.1 filename so v4.1 preferences persist.
        self._load_settings()
        self._apply_initial_input()
        self._annotate_dynamic_rows()
        self._apply_font()

    def _annotate_dynamic_rows(self):
        try:
            tab = self.tabs["settings"]
            src = next(w for w in tab.winfo_children() if isinstance(w, ttk.LabelFrame) and "数据源" in str(w.cget("text")))
            ttk.Label(src, text="显示分行时按当前可见内容自动调整列表行高：最多显示5行；超过5行时额外显示半行提示。", foreground="#555").grid(
                row=9, column=0, columnspan=4, sticky="w", pady=(3, 0))
        except Exception:
            pass

    # ------------------------------------------------------------------
    # Tree filtering, visible-only selection, and adaptive multiline display
    # ------------------------------------------------------------------
    def _tree(self, parent, columns, headings, widths, checkbox=False):
        tree = super()._tree(parent, columns, headings, widths, checkbox)
        seq = getattr(self, "_v41_tree_seq", 0) + 1; self._v41_tree_seq = seq
        style_name = f"QATree{seq}.Treeview"
        self.tree_registry[tree]["style_name"] = style_name
        self.tree_registry[tree]["disabled_iids"] = set()
        tree.configure(style=style_name)
        original_insert = tree.insert
        def insert_and_resize(*args, **kwargs):
            iid = original_insert(*args, **kwargs)
            try: self.after_idle(lambda t=tree: self._refresh_tree_height(t))
            except Exception: pass
            return iid
        tree.insert = insert_and_resize  # type: ignore[method-assign]
        return tree

    def _tree_context_menu(self, tree: ttk.Treeview, event):
        region = tree.identify_region(event.x, event.y)
        col_id = tree.identify_column(event.x); columns = tree["columns"]
        idx = int(col_id[1:]) - 1 if col_id.startswith("#") else -1
        col = columns[idx] if 0 <= idx < len(columns) else None
        menu = tk.Menu(self, tearoff=False)
        if region == "heading" and col:
            visible = tree.column(col, "width") > 0
            menu.add_command(label=("隐藏列：" if visible else "显示列：") + str(tree.heading(col, "text")),
                             command=lambda: self._set_column_visible(tree, col, not visible))
            menu.add_separator()
            menu.add_command(label="显示全部列", command=lambda: self._set_all_columns(tree, True))
            menu.add_command(label="恢复本页默认列", command=lambda: self._restore_page_columns(self.tree_registry[tree]["page"]))
            menu.add_separator()
            menu.add_command(label="筛选当前列…", command=lambda: self._filter_tree_column(tree, col))
            menu.add_command(label="清除表格筛选", command=lambda: self._clear_tree_filter(tree))
        else:
            iid = tree.identify_row(event.y)
            if iid:
                tree.selection_set(iid)
                if col: menu.add_command(label="复制当前单元格", command=lambda: self._copy_tree_cell(tree, iid, col))
                menu.add_command(label="复制选中行", command=lambda: self._copy_tree_selection(tree))
        if menu.index("end") is not None: menu.tk_popup(event.x_root, event.y_root)

    def _filter_tree_column(self, tree: ttk.Treeview, col):
        # One active filter at a time, as in v4, but finite-option columns now
        # use a checkbox list instead of textual contains matching.
        self._clear_tree_filter(tree)
        all_iids = list(tree.get_children(""))
        self.tree_registry[tree]["filter_order"] = list(all_iids)
        values = sorted({str(tree.set(iid, col)) for iid in all_iids}, key=lambda x: x.casefold())
        finite = (col in FIXED_FILTER_COLUMNS or
                  (col not in TEXT_FILTER_COLUMNS and 0 < len(values) <= 16 and max((len(v) for v in values), default=0) <= 45))
        meta = self.tree_registry[tree]
        detached: list[str] = []
        if finite:
            dlg = MultiChoiceFilterDialog(self, f"筛选：{tree.heading(col, 'text')}", values)
            self.wait_window(dlg)
            if dlg.result is None: return
            selected = dlg.result
            for iid in all_iids:
                if str(tree.set(iid, col)) not in selected:
                    tree.detach(iid); detached.append(iid)
        else:
            value = simpledialog.askstring("筛选当前列", "只显示包含以下文字的行；留空表示清除筛选：", parent=self)
            if value is None or not value: return
            needle = value.casefold()
            for iid in all_iids:
                if needle not in str(tree.set(iid, col)).casefold():
                    tree.detach(iid); detached.append(iid)
        meta["detached"] = detached
        self._refresh_tree_height(tree)

    def _clear_tree_filter(self, tree):
        meta = self.tree_registry.get(tree, {})
        order = list(meta.get("filter_order", []))
        super()._clear_tree_filter(tree)
        # Rebuild the exact pre-filter order (which may itself be a sorted order).
        for index, iid in enumerate(order):
            try:
                tree.move(iid, "", index)
            except Exception:
                pass
        meta["filter_order"] = []
        self._refresh_tree_height(tree)

    def _check_all(self, tree, checked=True, predicate=None):
        # get_children("") returns attached/visible rows only. Detached rows from
        # a filter are deliberately untouched.
        disabled = self.tree_registry.get(tree, {}).get("disabled_iids", set())
        for iid in tree.get_children(""):
            if iid in disabled: continue
            vals = list(tree.item(iid, "values"))
            if not vals: continue
            if predicate is None or predicate(iid, vals):
                vals[0] = "☑" if checked else "☐"; tree.item(iid, values=vals)

    def _checked_iids(self, tree):
        disabled = self.tree_registry.get(tree, {}).get("disabled_iids", set())
        return [iid for iid in tree.get_children("") if iid not in disabled and tree.item(iid, "values") and tree.item(iid, "values")[0] == "☑"]

    def _toggle_tree_checkbox(self, tree, event):
        iid = tree.identify_row(event.y)
        if iid in self.tree_registry.get(tree, {}).get("disabled_iids", set()): return "break"
        return super()._toggle_tree_checkbox(tree, event)

    def _display(self, text: str, limit: int = 420) -> str:
        value = (text or "").replace("\r\n", "\n").replace("\r", "\n")
        if self.display_line_mode.get() == "显示分行":
            lines = value.split("\n")
            if len(lines) > 5:
                value = "\n".join(lines[:5]) + f"\n↕ 还有 {len(lines)-5} 行"
            if limit is not None and len(value) > limit:
                value = value[:max(0, limit - 1)] + "…"
            return value
        value = value.replace("\n", " ↵ ")
        if limit is not None and len(value) > limit: value = value[:max(0, limit - 1)] + "…"
        return value

    def _apply_font(self):
        text_family = self.display_font.get() if hasattr(self, "display_font") else "系统默认"
        if text_family == "系统默认": text_family = tkfont.nametofont("TkTextFont").actual("family")
        list_family = self.list_font.get() if hasattr(self, "list_font") else "系统默认"
        if list_family == "系统默认": list_family = tkfont.nametofont("TkDefaultFont").actual("family")
        text_size = max(6, int(self.text_font_size.get() or 11)) if hasattr(self, "text_font_size") else 11
        list_size = max(6, int(self.list_font_size.get() or 10)) if hasattr(self, "list_font_size") else 10
        for widget in getattr(self, "text_widgets", []):
            try: widget.configure(font=(text_family, text_size))
            except Exception: pass
        if hasattr(self, "style"):
            self.style.configure("Treeview.Heading", font=(list_family, list_size, "bold"))
            for tree, meta in getattr(self, "tree_registry", {}).items():
                style_name = meta.get("style_name", "Treeview")
                self.style.configure(style_name, font=(list_family, list_size))
                self._refresh_tree_height(tree)
        try: self._draw_width_preview()
        except Exception: pass

    def _refresh_tree_height(self, tree):
        if not hasattr(self, "display_line_mode") or not hasattr(self, "style"): return
        meta = self.tree_registry.get(tree, {}); style_name = meta.get("style_name", "Treeview")
        size = max(6, int(self.list_font_size.get() or 10)) if hasattr(self, "list_font_size") else 10
        if self.display_line_mode.get() != "显示分行":
            self.style.configure(style_name, rowheight=max(25, size + 14)); return
        max_lines = 1
        for iid in tree.get_children(""):
            vals = tree.item(iid, "values")
            if vals:
                max_lines = max(max_lines, max(str(v).count("\n") + 1 for v in vals))
        units = min(max_lines, 5)
        if max_lines > 5: units += 0.5
        line_px = max(16, size + 8)
        self.style.configure(style_name, rowheight=max(28, int(math.ceil(line_px * units + 6))))

    def _toggle_line_mode(self):
        multiline = self.display_line_mode.get() == "显示分行"
        for tree in self.tree_registry:
            for iid in tree.get_children(""):
                vals = list(tree.item(iid, "values"))
                vals = [str(v).replace(" ↵ ", "\n") if multiline else str(v).replace("\n", " ↵ ") for v in vals]
                tree.item(iid, values=vals)
            self._refresh_tree_height(tree)
        self._apply_font()

    # ------------------------------------------------------------------
    # Editor: RM2k/2k3 operator reference and rendered hidden mode
    # ------------------------------------------------------------------
    def _build_editor(self):
        super()._build_editor()
        tab = self.tabs["editor"]
        extra = ttk.Frame(tab); extra.grid(row=5, column=0, sticky="ew", pady=(2, 0))
        ttk.Button(extra, text="RPG Maker 2000/2003 操作符一览", command=self._open_control_reference).pack(side="left")
        ttk.Button(extra, text="打开中文操作符文档", command=lambda: open_file(self.control_doc_path)).pack(side="left", padx=5)
        ttk.Label(extra, text="双击操作符可插入到译文光标位置；隐藏操作符时会显示可静态呈现的特殊符号。", foreground="#555").pack(side="left", padx=10)

    def _toggle_editor_controls(self):
        if not self.current_record: return
        if not self.editor_controls_hidden:
            self.editor_raw_translation = self.editor_translation.get("1.0", "end-1c")
            self._set_text(self.editor_original, render_rm2k3_controls(self.editor_raw_original), True)
            self._set_text(self.editor_translation, render_rm2k3_controls(self.editor_raw_translation), True)
            self.editor_controls_hidden = True; self.editor_control_button.configure(text="显示 RPG Maker 通配符/操作符")
        else:
            self._set_text(self.editor_original, self.editor_raw_original, True)
            self._set_text(self.editor_translation, self.editor_raw_translation, False)
            self.editor_controls_hidden = False; self.editor_control_button.configure(text="隐藏 RPG Maker 通配符/操作符")

    def _load_control_catalog(self):
        try: return json.loads(self.control_catalog_path.read_text(encoding="utf-8"))
        except Exception as exc:
            messagebox.showerror("操作符一览", f"无法读取：{self.control_catalog_path}\n{exc}", parent=self); return []

    def _control_usage_count(self, item: dict) -> int:
        if not self.current_record: return 0
        text = self.current_record.original + "\n" + self.editor_translation.get("1.0", "end-1c")
        try: return len(re.findall(item.get("regex", re.escape(item.get("insert", ""))), text, re.I))
        except re.error: return text.count(item.get("insert", ""))

    def _open_control_reference(self):
        rows = self._load_control_catalog()
        top = tk.Toplevel(self); top.title("RPG Maker 2000/2003 操作符一览"); top.geometry("1080x720")
        ttk.Label(top, text="以下默认表仅收录 RPG Maker 2000/2003 的消息控制符与 $A-$z 特殊符号。双击一行插入“插入形式”。", foreground="#444").pack(anchor="w", padx=8, pady=6)
        tree = ttk.Treeview(top, columns=("code", "insert", "meaning", "effect", "used"), show="headings")
        for c, title, width in [("code","操作符",150),("insert","插入形式",150),("meaning","作用",280),("effect","显示/运行效果",360),("used","当前文本使用",100)]:
            tree.heading(c, text=title); tree.column(c, width=width, anchor="w")
        y = ttk.Scrollbar(top, orient="vertical", command=tree.yview); tree.configure(yscrollcommand=y.set)
        tree.pack(side="left", fill="both", expand=True, padx=(8,0), pady=(0,8)); y.pack(side="right", fill="y", padx=(0,8), pady=(0,8))
        by_iid = {}
        for i, item in enumerate(rows):
            iid=f"op{i}"; by_iid[iid]=item
            tree.insert("", "end", iid=iid, values=(item.get("code",""), item.get("insert",""), item.get("meaning",""), item.get("effect",""), self._control_usage_count(item)))
        def insert_op(_e=None):
            sel=tree.selection()
            if not sel:return
            if self.editor_controls_hidden:
                messagebox.showinfo("插入操作符", "请先在单独文本处理页显示操作符，再插入。", parent=top); return
            value=str(by_iid[sel[0]].get("insert", ""))
            if value:
                self.editor_translation.insert("insert", value); self.editor_translation.focus_set()
        tree.bind("<Double-1>", insert_op)

    # ------------------------------------------------------------------
    # Quote analysis: one row per record with separate situation/style columns
    # ------------------------------------------------------------------
    def _build_quote_panel(self):
        tab = self.punct_tabs["quote"]; tab.rowconfigure(2, weight=1); tab.columnconfigure(0, weight=1)
        bar = ttk.Frame(tab); bar.grid(row=0, column=0, sticky="ew")
        ttk.Button(bar, text="提取全部含引号文本", command=self._do_quote_analysis).pack(side="left")
        ttk.Label(bar, text="引号状况处理：").pack(side="left", padx=(12,2))
        self.quote_form=tk.StringVar(value="与原文一致")
        ttk.Combobox(bar,textvariable=self.quote_form,state="readonly",values=["与原文一致","左右引号","只有左引号","删除引号"],width=12).pack(side="left")
        ttk.Label(bar,text="引号样式：").pack(side="left",padx=(8,2)); self.quote_style=tk.StringVar(value="与原文相同")
        ttk.Combobox(bar,textvariable=self.quote_style,state="readonly",values=["与原文相同"]+QUOTE_STYLE_LABELS,width=12).pack(side="left")
        ttk.Button(bar,text="生成处理预览",command=self._preview_quote_changes).pack(side="left",padx=5)
        ttk.Button(bar,text="应用勾选",command=self._apply_quote_changes).pack(side="right")
        act=ttk.Frame(tab); act.grid(row=1,column=0,sticky="ew",pady=4)
        ttk.Button(act,text="全选可批量项",command=lambda:self._check_all(self.quote_tree,True)).pack(side="left")
        ttk.Button(act,text="全部取消",command=lambda:self._check_all(self.quote_tree,False)).pack(side="left",padx=4)
        ttk.Button(act,text="只选状况不符",command=lambda:self._quote_select_mismatch("situation")).pack(side="left")
        ttk.Button(act,text="只选样式不符",command=lambda:self._quote_select_mismatch("style")).pack(side="left",padx=4)
        self.quote_tree=self._tree(tab,
            ("use","file","situation","situation_match","style","style_match","batch","original","translated","proposed"),
            ("选择","文件","引号状况","状况与原文相符","引号样式","样式与原文相符","批量处理能力","原文","译文","处理结果"),
            (60,200,360,120,300,120,180,400,400,400),checkbox=True)
        self.quote_tree.master.grid(row=2,column=0,sticky="nsew"); self.quote_tree.bind("<Double-1>",lambda e:self._double_to_editor("quote"))
        self.quote_rows: dict[str, CombinedQuoteAnalysisRow] = {}

    def _do_quote_analysis(self):
        try:
            rows=analyze_quote_combined(self.active_records()); self.quote_rows={}; self.result_maps["quote"]={}; self._clear_tree(self.quote_tree)
            disabled=set()
            for i,row in enumerate(rows):
                iid=f"q{i}"; self.quote_rows[iid]=row; self.result_maps["quote"][iid]=row.record
                batch=("状况可；" if row.situation_auto else "状况手动；")+("样式可" if row.style_auto else "样式手动")
                if not row.situation_auto and not row.style_auto: disabled.add(iid)
                self.quote_tree.insert("","end",iid=iid,values=("☐",row.record.file_key,row.situation,row.situation_match,row.style,row.style_match,batch,
                    self._display(row.record.original),self._display(row.record.translated),""))
            self.tree_registry[self.quote_tree]["disabled_iids"]=disabled
            self.status_var.set(f"提取 {len(rows)} 条含引号 Message；状况和样式已拆分为独立列。")
        except Exception as exc: messagebox.showerror("引号检查失败",str(exc),parent=self)

    def _quote_select_mismatch(self, which: str):
        for iid in self.quote_tree.get_children(""):
            row=self.quote_rows.get(iid); vals=list(self.quote_tree.item(iid,"values"))
            if not row:continue
            ok=(row.situation_match=="否" and row.situation_auto) if which=="situation" else (row.style_match=="否" and row.style_auto)
            vals[0]="☑" if ok else "☐"; self.quote_tree.item(iid,values=vals)

    def _preview_quote_changes(self):
        if not self.quote_rows:self._do_quote_analysis()
        for iid,row in list(self.quote_rows.items()):
            text=row.record.translated; changed=False
            temp=row.record
            if row.situation_auto:
                pair=None
                if self.quote_style.get()=="与原文相同" and len(row.source_styles)==1: pair=QUOTE_PAIRS[row.source_styles[0]]
                elif self.quote_style.get() in QUOTE_PAIRS: pair=QUOTE_PAIRS[self.quote_style.get()]
                form=self.quote_form.get()
                if form=="与原文一致":
                    info=_outer_quote_info(row.record.original); form="左右引号" if info.get("close_at_end") else "只有左引号"
                proposal=build_quote_situation_proposal(temp,form,pair)
                if proposal is not None: text=proposal; changed=changed or text!=temp.translated; temp=dataclass_replace(temp,translated=text)
            if row.style_auto:
                pair=None
                if self.quote_style.get()=="与原文相同" and len(row.source_styles)==1: pair=QUOTE_PAIRS[row.source_styles[0]]
                elif self.quote_style.get() in QUOTE_PAIRS: pair=QUOTE_PAIRS[self.quote_style.get()]
                if pair is not None:
                    text2=build_quote_style_proposal(temp,pair); changed=changed or text2!=text; text=text2
            proposal=text if changed else None
            self.quote_rows[iid]=dataclass_replace(row,proposal=proposal)
            vals=list(self.quote_tree.item(iid,"values")); vals[-1]=self._display(proposal or ""); self.quote_tree.item(iid,values=vals)

    def _apply_quote_changes(self):
        updates={}
        for iid in self._checked_iids(self.quote_tree):
            row=self.quote_rows.get(iid)
            if row and row.proposal is not None and row.proposal!=row.record.translated: updates[row.record.uid]=(row.record,row.proposal)
        self._apply_updates_dialog(updates,"引号处理")

    # ------------------------------------------------------------------
    # Ellipsis UI with clarified conversion language
    # ------------------------------------------------------------------
    def _build_ellipsis_panel(self):
        tab=self.punct_tabs["ellipsis"]; tab.rowconfigure(3,weight=1); tab.columnconfigure(0,weight=1)
        scan=ttk.Frame(tab); scan.grid(row=0,column=0,sticky="ew")
        self.ellipsis_side=tk.StringVar(value="both"); self.ellipsis_kind=tk.StringVar(value="全部")
        ttk.Label(scan,text="检查：").pack(side="left"); ttk.Combobox(scan,textvariable=self.ellipsis_side,state="readonly",values=["original","translated","both"],width=11).pack(side="left")
        ttk.Combobox(scan,textvariable=self.ellipsis_kind,state="readonly",values=["全部","连续点","省略号"],width=9).pack(side="left",padx=4)
        ttk.Button(scan,text="分类检查",command=self._do_ellipsis_scan).pack(side="left")
        ttk.Label(scan,text="连续点仅指至少两个连续的相同点；“…”始终属于省略号。",foreground="#555").pack(side="left",padx=12)
        conv=ttk.LabelFrame(tab,text="转换",padding=4); conv.grid(row=1,column=0,sticky="ew",pady=4)
        self.ellipsis_direction=tk.StringVar(value="连续点→省略号")
        ttk.Combobox(conv,textvariable=self.ellipsis_direction,state="readonly",values=["连续点→省略号","省略号→连续点","省略号→省略号"],width=16).grid(row=0,column=0,padx=3)
        ttk.Label(conv,text="连续点→省略号：每").grid(row=0,column=1); self.ellipsis_group_size=tk.IntVar(value=2)
        ttk.Spinbox(conv,from_=1,to=12,textvariable=self.ellipsis_group_size,width=4).grid(row=0,column=2); ttk.Label(conv,text="点对应一个省略号，多余点").grid(row=0,column=3)
        self.ellipsis_remainder=tk.StringVar(value="删除"); ttk.Combobox(conv,textvariable=self.ellipsis_remainder,state="readonly",values=["删除","一个省略号"],width=10).grid(row=0,column=4,padx=3)
        ttk.Label(conv,text="需要转换的省略号为：").grid(row=0,column=5); self.ellipsis_unit_style=tk.StringVar(value="双省略号")
        ttk.Combobox(conv,textvariable=self.ellipsis_unit_style,state="readonly",values=["单省略号","双省略号"],width=10).grid(row=0,column=6)
        ttk.Label(conv,text="省略号→连续点：每一个省略号对应").grid(row=1,column=1,pady=3); self.ellipsis_dots_per=tk.IntVar(value=3)
        ttk.Spinbox(conv,from_=1,to=12,textvariable=self.ellipsis_dots_per,width=4).grid(row=1,column=2); ttk.Label(conv,text="点，点符号").grid(row=1,column=3)
        self.ellipsis_dot_style=tk.StringVar(value="."); ttk.Combobox(conv,textvariable=self.ellipsis_dot_style,state="readonly",values=[".","．","·","・","。"],width=5).grid(row=1,column=4)
        ttk.Label(conv,text="省略号→省略号：统一为").grid(row=1,column=5); self.ellipsis_target_style=tk.StringVar(value="双省略号")
        ttk.Combobox(conv,textvariable=self.ellipsis_target_style,state="readonly",values=["单省略号","双省略号"],width=10).grid(row=1,column=6)
        self.ellipsis_source_vars={ch:tk.BooleanVar(value=True) for ch in [".","．","·","・","。"]}
        pf=ttk.Frame(conv); pf.grid(row=2,column=1,columnspan=6,sticky="w")
        ttk.Label(pf,text="允许转换的连续点：").pack(side="left")
        for ch,var in self.ellipsis_source_vars.items():ttk.Checkbutton(pf,text=ch,variable=var).pack(side="left")
        ttk.Button(pf,text="预览勾选转换",command=self._preview_ellipsis).pack(side="left",padx=8)
        ttk.Button(pf,text="应用勾选",command=lambda:self._apply_generic_changes("ellipsis_v4")).pack(side="left")
        act=ttk.Frame(tab); act.grid(row=2,column=0,sticky="ew")
        ttk.Button(act,text="全选可处理项",command=lambda:self._check_all(self.ellipsis_tree,True)).pack(side="left")
        ttk.Button(act,text="全不选",command=lambda:self._check_all(self.ellipsis_tree,False)).pack(side="left",padx=4)
        self.ellipsis_tree=self._tree(tab,("use","side","kind","style","count","control","file","manual","reason","original","translated","proposed"),
            ("选择","文本侧","类型","形式","数量","夹操作符","文件","仅手动","说明","原文","译文","处理结果"),
            (60,70,90,100,60,80,200,80,250,340,340,340),checkbox=True)
        self.ellipsis_tree.master.grid(row=3,column=0,sticky="nsew"); self.ellipsis_tree.bind("<Double-1>",lambda e:self._double_to_editor("ellipsis_v4"))
        self.ellipsis_occurrences={}

    def _do_ellipsis_scan(self):
        try:
            rows=analyze_ellipsis_occurrences(self.active_records(),self.ellipsis_side.get())
            if self.ellipsis_kind.get()!="全部":rows=[x for x in rows if x.kind==self.ellipsis_kind.get()]
            self.ellipsis_occurrences={};self.result_maps["ellipsis_v4"]={};self._clear_tree(self.ellipsis_tree);disabled=set()
            for i,row in enumerate(rows):
                iid=f"el{i}";self.ellipsis_occurrences[iid]=row;self.result_maps["ellipsis_v4"][iid]=row.record
                if row.manual_only or row.side!="译文":disabled.add(iid)
                self.ellipsis_tree.insert("","end",iid=iid,values=("☐",row.side,row.kind,row.visible_style,row.count,"是" if row.interrupted_by_control else "否",row.record.file_key,
                    "是" if row.manual_only else "否",row.reason,self._display(row.record.original),self._display(row.record.translated),""))
            self.tree_registry[self.ellipsis_tree]["disabled_iids"]=disabled
            self.status_var.set(f"省略号/连续点共 {len(rows)} 项；单个点不会列入连续点。")
        except Exception as exc:messagebox.showerror("省略号检查失败",str(exc),parent=self)

    def _preview_ellipsis(self):
        records={}
        for iid in self._checked_iids(self.ellipsis_tree):
            occ=self.ellipsis_occurrences.get(iid)
            if occ and not occ.manual_only and occ.side=="译文":records[occ.record.uid]=occ.record
        changes=[]; style="…" if self.ellipsis_unit_style.get()=="单省略号" else "……"; target_count=1 if self.ellipsis_target_style.get()=="单省略号" else 2
        allowed={ch for ch,var in self.ellipsis_source_vars.items() if var.get()}
        for record in records.values():
            proposed=build_ellipsis_conversion(record,direction=self.ellipsis_direction.get(),source_chars=allowed,group_size=self.ellipsis_group_size.get(),
                remainder=self.ellipsis_remainder.get(),ellipsis_style=style,dot_style=self.ellipsis_dot_style.get(),dots_per_ellipsis=self.ellipsis_dots_per.get(),target_ellipsis_count=target_count)
            if proposed and proposed!=record.translated:changes.append(TextChange(record,self.ellipsis_direction.get(),proposed))
        self.result_maps["ellipsis_v4_changes"]={x.record.uid:x for x in changes};self.result_maps["ellipsis_v4_apply"]={x.record.uid:x for x in changes}
        for iid,occ in self.ellipsis_occurrences.items():
            change=self.result_maps["ellipsis_v4_changes"].get(occ.record.uid);vals=list(self.ellipsis_tree.item(iid,"values"));vals[-1]=self._display(change.proposed) if change else "";self.ellipsis_tree.item(iid,values=vals)

    # ------------------------------------------------------------------
    # English case / character width: source-profile filters and match column
    # ------------------------------------------------------------------
    def _build_alnum_workspace(self):
        tab=self.tabs["alnum"];tab.rowconfigure(0,weight=1);tab.columnconfigure(0,weight=1)
        nb=ttk.Notebook(tab);nb.grid(row=0,column=0,sticky="nsew")
        case_tab=ttk.Frame(nb,padding=6);width_tab=ttk.Frame(nb,padding=6);nb.add(case_tab,text="英文大小写");nb.add(width_tab,text="字符全半角")
        self._build_case_panel(case_tab);self._build_width_char_panel(width_tab)

    def _build_case_panel(self, tab):
        tab.rowconfigure(2,weight=1);tab.columnconfigure(0,weight=1)
        top=ttk.Frame(tab);top.grid(row=0,column=0,sticky="ew")
        self.case_profile_filter=tk.StringVar(value="全部");self.case_query=tk.StringVar()
        ttk.Label(top,text="查询范围：").pack(side="left");ttk.Combobox(top,textvariable=self.case_profile_filter,state="readonly",values=["全部","原文大写","原文小写","原文首字母大写","原文大小写混搭"],width=18).pack(side="left")
        ttk.Label(top,text="仅处理包含：").pack(side="left",padx=(8,2));ttk.Entry(top,textvariable=self.case_query,width=24).pack(side="left")
        ttk.Button(top,text="扫描",command=lambda:self._scan_alnum("case",False)).pack(side="left",padx=4);ttk.Button(top,text="查询与原文格式不相符",command=lambda:self._scan_alnum("case",True)).pack(side="left")
        ttk.Button(top,text="全选",command=lambda:self._check_all(self.case_tree,True)).pack(side="left",padx=(8,2));ttk.Button(top,text="全不选",command=lambda:self._check_all(self.case_tree,False)).pack(side="left")
        edit=ttk.LabelFrame(tab,text="修改框：有选区只处理选区；没有选区处理全文",padding=4);edit.grid(row=1,column=0,sticky="ew",pady=5);edit.columnconfigure(0,weight=1)
        self.case_edit=tk.Text(edit,height=4,wrap="none");self.case_edit.grid(row=0,column=0,sticky="ew");self.text_widgets.append(self.case_edit)
        self.case_tree=self._tree(tab,("use","file","found","source_profile","translated_profile","match","original","translated","proposed"),
            ("选择","文件","匹配字符","原文大小写格式","译文大小写格式","与原文相符","原文","译文","处理结果"),(60,200,160,140,140,100,400,400,400),checkbox=True)
        self.case_tree.master.grid(row=2,column=0,sticky="nsew");self.case_tree.bind("<Double-1>",lambda e:self._double_to_editor("case"));self.case_tree.bind("<<TreeviewSelect>>",lambda e:self._alnum_select_to_edit("case"))
        bar=ttk.Frame(tab);bar.grid(row=3,column=0,sticky="ew",pady=4);self.case_mode=tk.StringVar(value="首字母大写")
        ttk.Combobox(bar,textvariable=self.case_mode,state="readonly",values=["大写","小写","首字母大写"],width=12).pack(side="left")
        ttk.Button(bar,text="处理修改框划取/全文",command=lambda:self._transform_alnum_edit("case",self.case_mode.get())).pack(side="left",padx=4)
        ttk.Button(bar,text="修改框内容送入选中记录预览",command=lambda:self._alnum_edit_to_selected("case")).pack(side="left")
        ttk.Button(bar,text="预览勾选记录",command=lambda:self._preview_alnum_records("case")).pack(side="left",padx=4);ttk.Button(bar,text="应用勾选",command=lambda:self._apply_generic_changes("case")).pack(side="right")

    def _build_width_char_panel(self, tab):
        tab.rowconfigure(2,weight=1);tab.columnconfigure(0,weight=1)
        top=ttk.Frame(tab);top.grid(row=0,column=0,sticky="ew")
        self.charwidth_target=tk.StringVar(value="both");self.charwidth_profile_filter=tk.StringVar(value="全部");self.charwidth_query=tk.StringVar()
        ttk.Label(top,text="字符范围：").pack(side="left");ttk.Combobox(top,textvariable=self.charwidth_target,state="readonly",values=["english","digits","both"],width=10).pack(side="left")
        ttk.Label(top,text="原文全半角：").pack(side="left",padx=(8,2));ttk.Combobox(top,textvariable=self.charwidth_profile_filter,state="readonly",values=["全部","原文半角","原文全角","原文混搭"],width=12).pack(side="left")
        ttk.Label(top,text="仅处理包含：").pack(side="left",padx=(8,2));ttk.Entry(top,textvariable=self.charwidth_query,width=20).pack(side="left")
        ttk.Button(top,text="扫描",command=lambda:self._scan_alnum("charwidth",False)).pack(side="left",padx=4);ttk.Button(top,text="查询与原文格式不相符",command=lambda:self._scan_alnum("charwidth",True)).pack(side="left")
        ttk.Button(top,text="全选",command=lambda:self._check_all(self.charwidth_tree,True)).pack(side="left",padx=(8,2));ttk.Button(top,text="全不选",command=lambda:self._check_all(self.charwidth_tree,False)).pack(side="left")
        edit=ttk.LabelFrame(tab,text="修改框：有选区只处理选区；没有选区处理全文",padding=4);edit.grid(row=1,column=0,sticky="ew",pady=5);edit.columnconfigure(0,weight=1)
        self.charwidth_edit=tk.Text(edit,height=4,wrap="none");self.charwidth_edit.grid(row=0,column=0,sticky="ew");self.text_widgets.append(self.charwidth_edit)
        self.charwidth_tree=self._tree(tab,("use","file","found","source_profile","translated_profile","match","original","translated","proposed"),
            ("选择","文件","匹配字符","原文全半角格式","译文全半角格式","与原文相符","原文","译文","处理结果"),(60,200,160,140,140,100,400,400,400),checkbox=True)
        self.charwidth_tree.master.grid(row=2,column=0,sticky="nsew");self.charwidth_tree.bind("<Double-1>",lambda e:self._double_to_editor("charwidth"));self.charwidth_tree.bind("<<TreeviewSelect>>",lambda e:self._alnum_select_to_edit("charwidth"))
        bar=ttk.Frame(tab);bar.grid(row=3,column=0,sticky="ew",pady=4);self.english_width_mode=tk.StringVar(value="不变");self.digit_width_mode=tk.StringVar(value="不变")
        ttk.Label(bar,text="英文：").pack(side="left");ttk.Combobox(bar,textvariable=self.english_width_mode,state="readonly",values=["不变","全角","半角"],width=8).pack(side="left")
        ttk.Label(bar,text="数字：").pack(side="left",padx=(8,2));ttk.Combobox(bar,textvariable=self.digit_width_mode,state="readonly",values=["不变","全角","半角"],width=8).pack(side="left")
        ttk.Button(bar,text="处理修改框划取/全文",command=self._transform_width_edit).pack(side="left",padx=4);ttk.Button(bar,text="修改框内容送入选中记录预览",command=lambda:self._alnum_edit_to_selected("charwidth")).pack(side="left")
        ttk.Button(bar,text="预览勾选记录",command=lambda:self._preview_alnum_records("charwidth")).pack(side="left",padx=4);ttk.Button(bar,text="应用勾选",command=lambda:self._apply_generic_changes("charwidth")).pack(side="right")

    def _scan_alnum(self, prefix, mismatch_only=False):
        try:
            target="english" if prefix=="case" else self.charwidth_target.get();rows=analyze_alnum_occurrences(self.active_records(),target)
            query=getattr(self,prefix+"_query").get(); tree=getattr(self,prefix+"_tree");self._clear_tree(tree);self.result_maps[prefix]={};kept=[]
            for row in rows:
                rec=row.record
                if query and query.casefold() not in row.reason.casefold():continue
                if prefix=="case":
                    sp=case_profile(rec.original);tp=case_profile(rec.translated).replace("原文","译文");match=case_format_matches(rec.original,rec.translated);flt=self.case_profile_filter.get()
                else:
                    sp=width_profile(rec.original,target);tp=width_profile(rec.translated,target).replace("原文","译文");match=width_format_matches(rec.original,rec.translated,target);flt=self.charwidth_profile_filter.get()
                if flt!="全部" and sp!=flt:continue
                if mismatch_only and match:continue
                kept.append((row,sp,tp,match))
            for i,(row,sp,tp,match) in enumerate(kept):
                iid=f"{prefix}{i}";self.result_maps[prefix][iid]=row.record
                tree.insert("","end",iid=iid,values=("☐",row.record.file_key,row.reason,sp,tp,"是" if match else "否",self._display(row.record.original),self._display(row.record.translated),""))
            self.status_var.set(f"{prefix} 扫描命中 {len(kept)} 条。")
        except Exception as exc:messagebox.showerror("扫描失败",str(exc),parent=self)

    # ------------------------------------------------------------------
    # Dictionary import: assign a whole-file category when category is absent
    # ------------------------------------------------------------------
    def _build_dictionary(self):
        super()._build_dictionary()
        tab=self.tabs["dictionary"]
        bar=ttk.Frame(tab);bar.grid(row=3,column=0,sticky="ew",pady=(4,0))
        ttk.Button(bar,text="全选当前可见行",command=lambda:self._select_visible_rows(self.dictionary_tree,True)).pack(side="left")
        ttk.Button(bar,text="全不选",command=lambda:self._select_visible_rows(self.dictionary_tree,False)).pack(side="left",padx=4)
        ttk.Label(bar,text="导入的辞典没有类别列时，会先询问整份辞典的默认类别。",foreground="#555").pack(side="left",padx=10)

    def _select_visible_rows(self, tree, selected=True):
        visible=list(tree.get_children(""))
        if selected: tree.selection_set(visible)
        else: tree.selection_remove(*visible)

    def _dictionary_has_category(self, path: Path) -> bool:
        if path.suffix.lower()==".csv":
            raw=path.read_bytes()
            for enc in ("utf-8-sig","utf-8","gb18030","cp932"):
                try:text=raw.decode(enc);break
                except UnicodeDecodeError:continue
            else:text=raw.decode("utf-8",errors="replace")
            row=next(csv.reader(text.splitlines()),[]);headers={str(x).strip().casefold() for x in row}
        else:
            wb=load_workbook(path,read_only=True,data_only=False);ws=wb.active;headers={str(c.value or "").strip().casefold() for c in next(ws.iter_rows(min_row=1,max_row=1))};wb.close()
        return bool(headers & {"类别","category","type"})

    def _load_dictionary_with_category(self, path: Path):
        rows=load_editable_dictionary(path)
        if self._dictionary_has_category(path):return rows
        dlg=CategoryDialog(self,path.name);self.wait_window(dlg)
        if dlg.result is None:raise QAError(f"已取消导入：{path.name}")
        return [EditableDictionaryRow(r.original,r.translation,dlg.result,r.source or path.name) for r in rows]

    def _import_dictionary(self):
        paths=filedialog.askopenfilenames(parent=self,filetypes=[("辞典","*.xlsx *.xlsm *.csv")])
        if not paths:return
        try:
            for p in paths:self.dictionary_rows.extend(self._load_dictionary_with_category(Path(p)))
            self._refresh_dictionary_tree();self.status_var.set(f"辞典页共 {len(self.dictionary_rows)} 行。")
        except Exception as exc:messagebox.showerror("导入失败",str(exc),parent=self)

    def _match_external_to_dictionary(self):
        paths=filedialog.askopenfilenames(parent=self,filetypes=[("辞典","*.xlsx *.xlsm *.csv")])
        if not paths:return
        try:
            rows=[]
            for p in paths:rows.extend(self._load_dictionary_with_category(Path(p)))
            used=build_used_dictionary(self.active_records(),rows);self.dictionary_rows.extend(used);self._refresh_dictionary_tree();self.status_var.set(f"匹配并加入 {len(used)} 行。")
        except Exception as exc:messagebox.showerror("匹配失败",str(exc),parent=self)

    def _do_dictcheck_external(self):
        paths=filedialog.askopenfilenames(parent=self,filetypes=[("辞典","*.xlsx *.xlsm *.csv")])
        if not paths:return
        try:
            rows=[]
            for p in paths:rows.extend(self._load_dictionary_with_category(Path(p)))
            entries=self._resolve_conflicts(rows)
            if entries is not None:self._show_dict_warnings(self._dictionary_warnings(entries))
        except Exception as exc:messagebox.showerror("检查失败",str(exc),parent=self)

    def _dictionary_warnings(self, entries):
        from workbench_core import analyze_dictionary
        return analyze_dictionary(self.active_records(),entries,self.dict_messages_only.get())


def main():
    parser = argparse.ArgumentParser(description="RPG制作大师校对工具")
    parser.add_argument("--initial-input")
    parser.add_argument("--initial-origin-dir")
    parser.add_argument("--initial-translated-dir")
    args = parser.parse_args()
    app=RPGMakerProofreadingApp(
        initial_input=args.initial_input,
        initial_origin_dir=args.initial_origin_dir,
        initial_translated_dir=args.initial_translated_dir,
    );app.mainloop()


if __name__=="__main__":main()
