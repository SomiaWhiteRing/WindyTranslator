from __future__ import annotations

import csv
import io
import json
import os
import re
import shutil
import subprocess
import sys
import unicodedata
import uuid
import zipfile
from collections import Counter, OrderedDict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path, PureWindowsPath
from typing import Callable, Iterable, Iterator, Mapping, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


APP_NAME = "RPG Maker 翻译多格式校对分析工具"
TRANSLATION_SHEET = "Translation"
BACKUP_DIRNAME = "_Backups"
NON_SPEAKER_IDS = {"", "NARRATION", "SYSTEM", "_ERASE_FACE_"}
DEFAULT_SPEAKER_ID = "NARRATION"
SYSTEM_TEXT_SPEAKER_ID = "SYSTEM"
ERASE_COMMAND_ID = "_ERASE_FACE_"
MULTILINE_BLOCK_MARKERS = {"Message", "StringPicture", "ScrollText"}

# WindyTranslator/RPG Maker 常见控制符。分析时忽略，写回时原样保留。
CONTROL_CODE_RE = re.compile(
    r"\\(?:[A-Za-z][A-Za-z0-9_]*(?:\[[^\]\r\n]*\]|<[^>\r\n]*>)?|[><!\.\|\^\$\\\{\}])"
)
RE_MARKER_LINE = re.compile(r"^\s*#([^#]+)#")
RE_IS_FACE_GRAPHIC_LINE = re.compile(
    r"^\s*\{{2,}.*?Select Face Graphic:.*?\}{2,}\s*$", re.IGNORECASE
)
RE_EXTRACT_FACE_GRAPHIC_CONTENT = re.compile(r"Select Face Graphic:(.*)", re.IGNORECASE)
RE_PAGE_SEPARATOR = re.compile(
    r"^(?:-{5,}Page\d+-{5,}|={5,}Page\d+={5,}|\*{5,}Entry\d+\*{5,})$",
    re.IGNORECASE,
)

QUOTE_PAIRS: dict[str, str] = {
    "“": "”",
    "‘": "’",
    "「": "」",
    "『": "』",
    "｢": "｣",
    "\"": "\"",
    "'": "'",
}
OPEN_QUOTES = set(QUOTE_PAIRS)
CLOSE_QUOTES = set(QUOTE_PAIRS.values())
ALL_QUOTES = OPEN_QUOTES | CLOSE_QUOTES
TERMINAL_PUNCTUATION = {"。", "；", ";", "！", "!", "？", "?"}

ORIGINAL_HEADERS = {"原文", "original"}
TRANSLATION_HEADERS = {"译文", "translation", "translated"}
MARKER_HEADERS = {"标记", "original_marker", "originalmarker", "marker"}
SPEAKER_HEADERS = {"说话人id", "speaker_id", "speakerid", "speaker"}

LogFn = Callable[[str], None]


class QAError(RuntimeError):
    pass


@dataclass(frozen=True)
class ParsedOccurrence:
    marker: str
    speaker_id: str
    text: str
    start_line: int
    end_line: int
    locator: str
    kind: str


@dataclass(frozen=True)
class DataRecord:
    uid: str
    source_kind: str  # json/txt/excel
    file_key: str
    original: str
    translated: str
    marker: str
    speaker_id: str
    original_location: str
    translated_location: str
    original_open_path: Path
    translated_open_path: Path
    update_ref: tuple
    metadata: Mapping[str, object] = field(default_factory=dict)

    def display_location(self, txt_position_mode: str = "translated") -> str:
        if self.source_kind != "txt":
            return self.translated_location
        return self.original_location if txt_position_mode == "original" else self.translated_location

    def display_open_path(self, txt_position_mode: str = "translated") -> Path:
        if self.source_kind != "txt":
            return self.translated_open_path
        return self.original_open_path if txt_position_mode == "original" else self.translated_open_path


@dataclass(frozen=True)
class SourceLoadResult:
    records: tuple[DataRecord, ...]
    warnings: tuple[str, ...]
    source_kind: str


@dataclass(frozen=True)
class LengthWarning:
    record: DataRecord
    line_number: int
    width_units: float
    visible_line: str

    @property
    def key(self) -> str:
        return f"{self.record.uid}|{self.line_number}"


@dataclass(frozen=True)
class QuotePair:
    open_char: str
    close_char: str
    open_pos: int
    close_pos: int


@dataclass(frozen=True)
class QuoteProfile:
    case_code: str  # case1..case5/none/irregular
    description: str
    open_char: str | None = None
    close_char: str | None = None
    open_pos: int | None = None
    close_pos: int | None = None
    open_line: int | None = None
    open_at_line_start: bool = False
    close_at_text_end: bool = False
    pairs: tuple[QuotePair, ...] = ()
    unmatched_opens: tuple[tuple[str, int], ...] = ()
    unmatched_closes: tuple[tuple[str, int], ...] = ()


@dataclass(frozen=True)
class DialogueIssue:
    issue_id: str
    record: DataRecord
    case_code: str
    category_label: str
    reasons: tuple[str, ...]
    current_translation: str
    proposed_translation: str | None
    auto_fixable: bool
    original_profile: QuoteProfile


@dataclass(frozen=True)
class DictionaryEntry:
    original: str
    translation: str
    source_file: Path


@dataclass(frozen=True)
class DictionaryWarning:
    record: DataRecord
    original_term: str
    expected_translation: str
    source_dictionary: Path
    original_visible: str
    translated_visible: str

    @property
    def key(self) -> str:
        return f"{self.record.uid}|{self.original_term}|{self.expected_translation}"


@dataclass(frozen=True)
class SearchHit:
    record: DataRecord
    matched_field: str  # 原文/译文/原文和译文
    original_excerpt: str
    translated_excerpt: str

    @property
    def key(self) -> str:
        return f"{self.record.uid}|{self.matched_field}"


@dataclass(frozen=True)
class GeneratedDictionaryRow:
    original: str
    translation: str
    source_dictionary: str
    match_count: int


def _noop_log(_: str) -> None:
    return None


def normalize_newlines(text: str) -> str:
    return (text or "").replace("\r\n", "\n").replace("\r", "\n")


def strip_control_codes(text: str) -> str:
    return CONTROL_CODE_RE.sub("", text or "")


def is_effective_speaker(speaker_id: str | None) -> bool:
    return (speaker_id or "").strip().upper() not in NON_SPEAKER_IDS


def fullwidth_units(text: str) -> float:
    half_units = 0
    for char in text:
        if char in "\r\n":
            continue
        category = unicodedata.category(char)
        if category in {"Mn", "Me", "Cf", "Cc"}:
            continue
        if char == "\t":
            half_units += 4
            continue
        half_units += 2 if unicodedata.east_asian_width(char) in {"W", "F", "A"} else 1
    return half_units / 2.0


def _safe_windows_file_key(file_key: str) -> str:
    candidate = PureWindowsPath(str(file_key).replace("/", "\\"))
    if candidate.is_absolute() or candidate.drive or candidate.root:
        raise QAError(f"TXT 相对路径不能是绝对路径：{file_key}")
    if any(part in {"", ".", ".."} for part in candidate.parts):
        raise QAError(f"TXT 相对路径不安全：{file_key}")
    return str(candidate)


def _native_from_file_key(file_key: str) -> Path:
    return Path(*PureWindowsPath(_safe_windows_file_key(file_key)).parts)


def _read_text_file(path: Path) -> tuple[str, bool, str]:
    raw = path.read_bytes()
    has_bom = raw.startswith(b"\xef\xbb\xbf")
    if has_bom:
        raw = raw[3:]
    text = raw.decode("utf-8", errors="replace")
    newline = "\r\n" if "\r\n" in text else "\n"
    return normalize_newlines(text), has_bom, newline


def _write_text_file(path: Path, text: str, *, has_bom: bool, newline: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    payload_text = normalize_newlines(text)
    if newline != "\n":
        payload_text = payload_text.replace("\n", newline)
    payload = payload_text.encode("utf-8")
    if has_bom:
        payload = b"\xef\xbb\xbf" + payload
    path.write_bytes(payload)


def _parse_face_graphic(details: str) -> str | None:
    details = details.strip()
    if details.lower() == "erase":
        return ERASE_COMMAND_ID
    parts = [part.strip() for part in details.split(",")]
    raw_name = parts[0] if parts else ""
    if not raw_name:
        return ERASE_COMMAND_ID
    name = raw_name.strip("'\"")
    if name.lower() == "erase":
        return ERASE_COMMAND_ID
    if len(parts) > 1 and parts[1].isdigit():
        return f"{name}_{parts[1]}"
    return name or None


def parse_string_script(path: Path) -> list[ParsedOccurrence]:
    text, _, _ = _read_text_file(path)
    lines = text.splitlines(keepends=True)
    occurrences: list[ParsedOccurrence] = []
    current_speaker = DEFAULT_SPEAKER_ID
    marker_counts: Counter[str] = Counter()
    i = 0
    while i < len(lines):
        stripped = lines[i].strip()
        if RE_PAGE_SEPARATOR.match(stripped):
            current_speaker = DEFAULT_SPEAKER_ID
            i += 1
            continue
        if RE_IS_FACE_GRAPHIC_LINE.match(stripped):
            match = RE_EXTRACT_FACE_GRAPHIC_CONTENT.search(stripped)
            if match:
                cleaned = match.group(1).split("}", 1)[0].strip()
                parsed = _parse_face_graphic(cleaned)
                if parsed == ERASE_COMMAND_ID:
                    current_speaker = DEFAULT_SPEAKER_ID
                elif parsed:
                    current_speaker = parsed
            i += 1
            continue
        marker_match = RE_MARKER_LINE.match(stripped)
        if not marker_match:
            i += 1
            continue
        marker = marker_match.group(1)
        marker_counts[marker] += 1
        instance = marker_counts[marker]
        speaker = current_speaker if marker == "Message" else SYSTEM_TEXT_SPEAKER_ID
        i += 1
        if marker in MULTILINE_BLOCK_MARKERS or marker.startswith("PluginCommand_"):
            start = i
            block: list[str] = []
            while i < len(lines) and lines[i].strip() != "##":
                block.append(lines[i])
                i += 1
            occurrences.append(
                ParsedOccurrence(
                    marker=marker,
                    speaker_id=speaker,
                    text="".join(block).rstrip("\n"),
                    start_line=start,
                    end_line=i,
                    locator=f"{marker}:{instance}:1",
                    kind="multiline",
                )
            )
            if i < len(lines) and lines[i].strip() == "##":
                i += 1
            continue
        if marker == "EventName":
            if i < len(lines):
                i += 1
            continue
        if marker == "Choice":
            sub = 0
            while i < len(lines) and lines[i].strip() != "##":
                sub += 1
                occurrences.append(
                    ParsedOccurrence(
                        marker=marker,
                        speaker_id=speaker,
                        text=lines[i].rstrip("\n").strip(),
                        start_line=i,
                        end_line=i + 1,
                        locator=f"{marker}:{instance}:{sub}",
                        kind="choice",
                    )
                )
                i += 1
            if i < len(lines) and lines[i].strip() == "##":
                i += 1
            continue
        if i < len(lines):
            occurrences.append(
                ParsedOccurrence(
                    marker=marker,
                    speaker_id=SYSTEM_TEXT_SPEAKER_ID,
                    text=lines[i].rstrip("\n").strip(),
                    start_line=i,
                    end_line=i + 1,
                    locator=f"{marker}:{instance}:1",
                    kind="single",
                )
            )
            i += 1
    return occurrences


def _line_range_text(path: Path, occurrence: ParsedOccurrence | None) -> str:
    if occurrence is None:
        return f"{path}（未找到对应条目）"
    start = occurrence.start_line + 1
    end = max(start, occurrence.end_line)
    return f"{path}（第 {start}–{end} 行；{occurrence.locator}）"


def _list_txt_files(root: Path) -> OrderedDict[str, Path]:
    if not root.is_dir():
        raise QAError(f"TXT 文件夹不存在：{root}")
    result: OrderedDict[str, Path] = OrderedDict()
    for path in sorted(root.rglob("*.txt"), key=lambda p: str(p).casefold()):
        key = str(PureWindowsPath(*path.relative_to(root).parts))
        result[key] = path
    return result


def load_txt_records(origin_dir: Path, translated_dir: Path, log: LogFn = _noop_log) -> SourceLoadResult:
    origin_files = _list_txt_files(origin_dir)
    translated_files = _list_txt_files(translated_dir)
    records: list[DataRecord] = []
    warnings: list[str] = []
    for idx, (file_key, original_path) in enumerate(origin_files.items(), start=1):
        log(f"读取 TXT [{idx}/{len(origin_files)}]：{file_key}")
        translated_path = translated_files.get(file_key)
        original_occ = parse_string_script(original_path)
        translated_occ = parse_string_script(translated_path) if translated_path and translated_path.exists() else []
        translated_by_locator = {occ.locator: occ for occ in translated_occ}
        if translated_path is None:
            translated_path = translated_dir / _native_from_file_key(file_key)
            warnings.append(f"译文 TXT 缺失，将以原文作为当前译文：{file_key}")
        for occurrence in original_occ:
            if occurrence.text == "":
                continue
            matched = translated_by_locator.get(occurrence.locator)
            translated_text = matched.text if matched is not None else occurrence.text
            uid = f"txt|{file_key}|{occurrence.locator}"
            records.append(
                DataRecord(
                    uid=uid,
                    source_kind="txt",
                    file_key=file_key,
                    original=occurrence.text,
                    translated=translated_text,
                    marker=occurrence.marker,
                    speaker_id=occurrence.speaker_id,
                    original_location=_line_range_text(original_path, occurrence),
                    translated_location=_line_range_text(translated_path, matched),
                    original_open_path=original_path,
                    translated_open_path=translated_path,
                    update_ref=(translated_path, occurrence.locator, occurrence.kind),
                    metadata={"original_occurrence": occurrence, "translated_occurrence": matched},
                )
            )
    for extra in sorted(set(translated_files) - set(origin_files)):
        warnings.append(f"译文文件夹中存在原文文件夹没有的 TXT，已忽略：{extra}")
    return SourceLoadResult(tuple(records), tuple(warnings), "txt")


def _no_duplicate_object(pairs):
    obj = OrderedDict()
    for key, value in pairs:
        if key in obj:
            raise QAError(f"JSON 中存在重复键：{key!r}")
        obj[key] = value
    return obj


def load_json_records(json_path: Path, log: LogFn = _noop_log) -> SourceLoadResult:
    if not json_path.is_file():
        raise QAError(f"JSON 文件不存在：{json_path}")
    try:
        with json_path.open("r", encoding="utf-8-sig") as fh:
            data = json.load(fh, object_pairs_hook=_no_duplicate_object)
    except json.JSONDecodeError as exc:
        raise QAError(f"JSON 语法错误：第 {exc.lineno} 行，第 {exc.colno} 列：{exc.msg}") from exc
    if not isinstance(data, dict):
        raise QAError("JSON 顶层必须是对象。")
    records: list[DataRecord] = []
    for file_index, (file_key, entries) in enumerate(data.items(), start=1):
        log(f"读取 JSON 文件项 [{file_index}/{len(data)}]：{file_key}")
        if not isinstance(entries, dict):
            raise QAError(f"JSON 文件项不是对象：{file_key}")
        raw_file_key = str(file_key)
        safe_key = _safe_windows_file_key(raw_file_key)
        for entry_index, (original, meta) in enumerate(entries.items(), start=1):
            if not isinstance(original, str) or not isinstance(meta, dict):
                raise QAError(f"JSON 记录格式错误：{safe_key} / 第 {entry_index} 条")
            translated = meta.get("text")
            if not isinstance(translated, str):
                raise QAError(f"JSON 记录缺少字符串 text：{safe_key} / {original[:40]!r}")
            marker = str(meta.get("original_marker") or "")
            speaker = str(meta.get("speaker_id") or "")
            uid = f"json|{safe_key}|{entry_index}"
            base = f"{json_path} :: {safe_key} :: 第 {entry_index} 条"
            records.append(
                DataRecord(
                    uid=uid,
                    source_kind="json",
                    file_key=safe_key,
                    original=original,
                    translated=translated,
                    marker=marker,
                    speaker_id=speaker,
                    original_location=f"{base} / 原文键",
                    translated_location=f"{base} / text",
                    original_open_path=json_path,
                    translated_open_path=json_path,
                    update_ref=(raw_file_key, original),
                    metadata={"entry_index": entry_index},
                )
            )
    return SourceLoadResult(tuple(records), tuple(), "json")


def _normalize_header(value: object) -> str:
    return str(value or "").strip().casefold().replace(" ", "")


def _find_header(headers: Mapping[str, int], candidates: set[str]) -> int | None:
    normalized_candidates = {_normalize_header(v) for v in candidates}
    for name, idx in headers.items():
        if _normalize_header(name) in normalized_candidates:
            return idx
    return None


def _iter_workbooks(table_dir: Path) -> Iterator[Path]:
    if not table_dir.is_dir():
        raise QAError(f"Excel 文件夹不存在：{table_dir}")
    for path in sorted(table_dir.rglob("*.xlsx"), key=lambda p: str(p).casefold()):
        if BACKUP_DIRNAME in path.parts or path.name.startswith("~$") or path.name.startswith("."):
            continue
        yield path
    for path in sorted(table_dir.rglob("*.xlsm"), key=lambda p: str(p).casefold()):
        if BACKUP_DIRNAME in path.parts or path.name.startswith("~$") or path.name.startswith("."):
            continue
        yield path


def load_excel_records(table_dir: Path, log: LogFn = _noop_log) -> SourceLoadResult:
    workbooks = list(_iter_workbooks(table_dir))
    if not workbooks:
        raise QAError(f"Excel 文件夹中没有 .xlsx/.xlsm 文件：{table_dir}")
    records: list[DataRecord] = []
    warnings: list[str] = []
    for idx, path in enumerate(workbooks, start=1):
        relative = str(path.relative_to(table_dir))
        log(f"读取 Excel [{idx}/{len(workbooks)}]：{relative}")
        try:
            wb = load_workbook(path, read_only=True, data_only=False)
        except Exception as exc:
            warnings.append(f"无法读取，已跳过：{relative}（{exc}）")
            continue
        try:
            candidate_sheets = [wb[TRANSLATION_SHEET]] if TRANSLATION_SHEET in wb.sheetnames else list(wb.worksheets)
            found = False
            for ws in candidate_sheets:
                max_row, _max_col = _worksheet_dimensions(ws)
                if max_row < 1:
                    continue
                values = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                headers = {str(value).strip(): col for col, value in enumerate(values, start=1) if value is not None}
                original_col = _find_header(headers, ORIGINAL_HEADERS)
                translation_col = _find_header(headers, TRANSLATION_HEADERS)
                if original_col is None or translation_col is None:
                    continue
                marker_col = _find_header(headers, MARKER_HEADERS)
                speaker_col = _find_header(headers, SPEAKER_HEADERS)
                found = True
                for row_num, row_values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    original_value = row_values[original_col - 1] if len(row_values) >= original_col else None
                    if original_value in (None, ""):
                        continue
                    translated_value = row_values[translation_col - 1] if len(row_values) >= translation_col else None
                    marker_value = row_values[marker_col - 1] if marker_col and len(row_values) >= marker_col else ""
                    speaker_value = row_values[speaker_col - 1] if speaker_col and len(row_values) >= speaker_col else ""
                    file_key = relative
                    base = f"{path} / {ws.title}"
                    uid = f"excel|{relative}|{ws.title}|{row_num}"
                    records.append(
                        DataRecord(
                            uid=uid,
                            source_kind="excel",
                            file_key=file_key,
                            original=str(original_value),
                            translated=str(original_value if translated_value is None else translated_value),
                            marker=str(marker_value or ""),
                            speaker_id=str(speaker_value or ""),
                            original_location=f"{base}!{_column_letter(original_col)}{row_num}",
                            translated_location=f"{base}!{_column_letter(translation_col)}{row_num}",
                            original_open_path=path,
                            translated_open_path=path,
                            update_ref=(path, ws.title, row_num, translation_col),
                            metadata={"table_dir": table_dir},
                        )
                    )
                break
            if not found:
                warnings.append(f"未找到原文/译文列，已跳过：{relative}")
        finally:
            wb.close()
    if not records:
        raise QAError("没有找到可分析的 Excel 翻译记录。")
    return SourceLoadResult(tuple(records), tuple(warnings), "excel")


def _worksheet_dimensions(ws) -> tuple[int, int]:
    if ws.max_row is None or ws.max_column is None:
        try:
            ws.calculate_dimension(force=True)
        except TypeError:
            ws.calculate_dimension()
    return int(ws.max_row or 0), int(ws.max_column or 0)


def _column_letter(index: int) -> str:
    chars: list[str] = []
    while index:
        index, rem = divmod(index - 1, 26)
        chars.append(chr(65 + rem))
    return "".join(reversed(chars))


def analyze_length(records: Iterable[DataRecord], limit: float = 19.0) -> list[LengthWarning]:
    results: list[LengthWarning] = []
    for record in records:
        if record.marker != "Message" or not is_effective_speaker(record.speaker_id):
            continue
        for line_num, line in enumerate(normalize_newlines(record.translated).split("\n"), start=1):
            visible = strip_control_codes(line)
            width = fullwidth_units(visible)
            if width > limit + 1e-9:
                results.append(LengthWarning(record, line_num, width, visible))
    return results


def _control_mask(text: str) -> list[bool]:
    mask = [True] * len(text)
    for match in CONTROL_CODE_RE.finditer(text):
        for i in range(match.start(), match.end()):
            mask[i] = False
    return mask


def _visible_positions(text: str, include_whitespace: bool = False) -> list[int]:
    mask = _control_mask(text)
    result: list[int] = []
    for idx, char in enumerate(text):
        if not mask[idx]:
            continue
        if not include_whitespace and char.isspace():
            continue
        result.append(idx)
    return result


def _line_first_visible_positions(text: str) -> dict[int, int]:
    mask = _control_mask(text)
    result: dict[int, int] = {}
    line_no = 0
    for idx, char in enumerate(text):
        if char == "\n":
            line_no += 1
            continue
        if not mask[idx] or char.isspace():
            continue
        result.setdefault(line_no, idx)
    return result


def _line_number_at(text: str, raw_pos: int) -> int:
    return text.count("\n", 0, raw_pos)


def _scan_quote_pairs(text: str) -> tuple[list[QuotePair], list[tuple[str, int]], list[tuple[str, int]]]:
    visible = _visible_positions(text)
    pairs: list[QuotePair] = []
    unmatched_opens: list[tuple[str, int]] = []
    unmatched_closes: list[tuple[str, int]] = []
    stacks: dict[str, list[int]] = {open_q: [] for open_q in QUOTE_PAIRS if open_q != QUOTE_PAIRS[open_q]}
    symmetric_stacks: dict[str, list[int]] = {q: [] for q in QUOTE_PAIRS if q == QUOTE_PAIRS[q]}
    close_to_open = {close_q: open_q for open_q, close_q in QUOTE_PAIRS.items() if open_q != close_q}
    for pos in visible:
        char = text[pos]
        if char in symmetric_stacks:
            stack = symmetric_stacks[char]
            if stack:
                open_pos = stack.pop()
                pairs.append(QuotePair(char, char, open_pos, pos))
            else:
                stack.append(pos)
            continue
        if char in stacks:
            stacks[char].append(pos)
            continue
        if char in close_to_open:
            open_q = close_to_open[char]
            stack = stacks[open_q]
            if stack:
                open_pos = stack.pop()
                pairs.append(QuotePair(open_q, char, open_pos, pos))
            else:
                unmatched_closes.append((char, pos))
    for open_q, stack in stacks.items():
        unmatched_opens.extend((open_q, pos) for pos in stack)
    for quote, stack in symmetric_stacks.items():
        unmatched_opens.extend((quote, pos) for pos in stack)
    pairs.sort(key=lambda p: (p.open_pos, p.close_pos))
    unmatched_opens.sort(key=lambda item: item[1])
    unmatched_closes.sort(key=lambda item: item[1])
    return pairs, unmatched_opens, unmatched_closes


def classify_quote_profile(text: str) -> QuoteProfile:
    pairs, unmatched_opens, unmatched_closes = _scan_quote_pairs(text)
    visible = _visible_positions(text)
    if not visible:
        return QuoteProfile("none", "无可见文本")
    last_visible = visible[-1]
    line_starts = _line_first_visible_positions(text)
    line_start_positions = set(line_starts.values())

    # Boundary/outer structures take priority over nested internal pairs.
    case1_candidates = [p for p in pairs if p.open_pos in line_start_positions and p.close_pos == last_visible]
    if case1_candidates:
        p = max(case1_candidates, key=lambda item: item.close_pos - item.open_pos)
        return QuoteProfile(
            "case1", "情形1：左引号在某行开头，配对右引号在全文结尾",
            p.open_char, p.close_char, p.open_pos, p.close_pos, _line_number_at(text, p.open_pos), True, True,
            tuple(pairs), tuple(unmatched_opens), tuple(unmatched_closes),
        )

    # An unmatched line-start opening quote represents case 4 even when the text
    # also contains a nested, correctly paired quotation.
    line_start_unmatched = [(q, pos) for q, pos in unmatched_opens if pos in line_start_positions]
    if line_start_unmatched:
        q, pos = line_start_unmatched[0]
        return QuoteProfile(
            "case4", "情形4：左引号在某行开头，但没有配对右引号",
            q, QUOTE_PAIRS[q], pos, None, _line_number_at(text, pos), True, False,
            tuple(pairs), tuple(unmatched_opens), tuple(unmatched_closes),
        )

    case2_candidates = [p for p in pairs if p.close_pos == last_visible]
    if case2_candidates:
        p = max(case2_candidates, key=lambda item: item.close_pos - item.open_pos)
        return QuoteProfile(
            "case2", "情形2：左引号位于文本中间，配对右引号在全文结尾",
            p.open_char, p.close_char, p.open_pos, p.close_pos, _line_number_at(text, p.open_pos), False, True,
            tuple(pairs), tuple(unmatched_opens), tuple(unmatched_closes),
        )

    # An unmatched closing quote at the boundary represents case 5 even when
    # independent nested quote pairs exist earlier in the text.
    end_unmatched = [(q, pos) for q, pos in unmatched_closes if pos == last_visible]
    if end_unmatched:
        q, pos = end_unmatched[-1]
        open_q = next((o for o, c in QUOTE_PAIRS.items() if c == q), None)
        return QuoteProfile(
            "case5", "情形5：全文结尾有右引号，但没有配对左引号",
            open_q, q, None, pos, None, False, True,
            tuple(pairs), tuple(unmatched_opens), tuple(unmatched_closes),
        )

    if pairs:
        p = max(pairs, key=lambda item: item.close_pos - item.open_pos)
        return QuoteProfile(
            "case3", "情形3：左右引号配对，但右引号不在全文结尾",
            p.open_char, p.close_char, p.open_pos, p.close_pos, _line_number_at(text, p.open_pos), p.open_pos in line_start_positions, False,
            tuple(pairs), tuple(unmatched_opens), tuple(unmatched_closes),
        )
    if any(text[pos] in ALL_QUOTES for pos in visible):
        return QuoteProfile(
            "irregular", "引号结构不属于五种可安全处理情形",
            pairs=tuple(pairs), unmatched_opens=tuple(unmatched_opens), unmatched_closes=tuple(unmatched_closes),
        )
    return QuoteProfile("none", "原文无引号")


def _replace_char(text: str, pos: int, char: str) -> str:
    return text[:pos] + char + text[pos + 1 :]


def _delete_char(text: str, pos: int) -> str:
    return text[:pos] + text[pos + 1 :]


def _insert_char(text: str, pos: int, char: str) -> str:
    return text[:pos] + char + text[pos:]


def _first_visible_on_line(text: str, line_no: int) -> int | None:
    return _line_first_visible_positions(text).get(line_no)


def _last_visible(text: str) -> int | None:
    positions = _visible_positions(text)
    return positions[-1] if positions else None


def _terminal_punctuation(text: str) -> tuple[str | None, int | None]:
    visible = _visible_positions(text)
    if not visible:
        return None, None
    idx = len(visible) - 1
    # 抛开结尾引号检查标点。
    while idx >= 0 and text[visible[idx]] in CLOSE_QUOTES:
        idx -= 1
    if idx < 0:
        return None, None
    raw_pos = visible[idx]
    char = text[raw_pos]
    if char in TERMINAL_PUNCTUATION:
        return char, raw_pos
    return None, None


def _terminal_raw_char_before_quotes(text: str) -> tuple[str | None, int | None]:
    visible = _visible_positions(text)
    if not visible:
        return None, None
    idx = len(visible) - 1
    while idx >= 0 and text[visible[idx]] in CLOSE_QUOTES:
        idx -= 1
    if idx < 0:
        return None, None
    raw = visible[idx]
    return text[raw], raw


def _insert_terminal_punctuation(text: str, punctuation: str) -> str:
    visible = _visible_positions(text)
    if not visible:
        return text
    idx = len(visible) - 1
    while idx >= 0 and text[visible[idx]] in CLOSE_QUOTES:
        idx -= 1
    if idx < 0:
        return text
    raw_pos = visible[idx]
    raw_char = text[raw_pos]
    if raw_char == ".":
        return _replace_char(text, raw_pos, punctuation)
    return _insert_char(text, raw_pos + 1, punctuation)


def _apply_punctuation_rule(original: str, translated: str) -> tuple[str, list[str]]:
    original_punct, _ = _terminal_punctuation(original)
    translated_punct, translated_pos = _terminal_punctuation(translated)
    reasons: list[str] = []
    result = translated
    if original_punct is None and translated_punct is not None and translated_pos is not None:
        reasons.append(f"结尾标点：原文无指定收尾标点，译文多出 {translated_punct!r}。")
        result = _delete_char(result, translated_pos)
    elif original_punct is not None:
        if translated_punct is None:
            reasons.append(f"结尾标点：原文以 {original_punct!r} 收尾，译文缺少。")
            result = _insert_terminal_punctuation(result, original_punct)
        elif translated_punct != original_punct and translated_pos is not None:
            reasons.append(
                f"结尾标点：原文为 {original_punct!r}，译文为 {translated_punct!r}，符号不一致。"
            )
            result = _replace_char(result, translated_pos, original_punct)
    return result, reasons


def _profile_translation_quotes(text: str) -> QuoteProfile:
    return classify_quote_profile(text)


def _ensure_line_start_open(
    text: str,
    line_no: int,
    target_open: str,
) -> tuple[str, list[str], bool]:
    reasons: list[str] = []
    line_start = _first_visible_on_line(text, line_no)
    if line_start is None:
        return text, [f"译文没有原文左引号所在的第 {line_no + 1} 行，需手动处理。"], False
    current = text[line_start]
    opens = [(q, pos) for q, pos in _scan_quote_pairs(text)[1]]
    opens += [(p.open_char, p.open_pos) for p in _scan_quote_pairs(text)[0]]
    opens = sorted(set(opens), key=lambda item: item[1])
    if current in OPEN_QUOTES:
        if current != target_open:
            reasons.append(f"左引号符号应由 {current!r} 改为 {target_open!r}。")
            return _replace_char(text, line_start, target_open), reasons, True
        return text, reasons, True
    if opens:
        reasons.append("译文存在左引号，但位置不在原文对应行的开头，需手动处理。")
        return text, reasons, False
    reasons.append(f"译文缺少位于第 {line_no + 1} 行开头的左引号 {target_open!r}。")
    return _insert_char(text, line_start, target_open), reasons, True


def _ensure_end_close(text: str, target_close: str) -> tuple[str, list[str], bool]:
    reasons: list[str] = []
    last = _last_visible(text)
    if last is None:
        return text, ["译文为空，无法补齐右引号。"], False
    current = text[last]
    if current in CLOSE_QUOTES:
        if current != target_close:
            reasons.append(f"右引号符号应由 {current!r} 改为 {target_close!r}。")
            return _replace_char(text, last, target_close), reasons, True
        return text, reasons, True
    reasons.append(f"译文结尾缺少右引号 {target_close!r}。")
    return _insert_char(text, last + 1, target_close), reasons, True


def _ensure_case1_end_close(text: str, target_close: str) -> tuple[str, list[str], bool]:
    last = _last_visible(text)
    if last is None:
        return text, ["译文为空，无法补齐右引号。"], False
    profile = _profile_translation_quotes(text)
    all_close_positions = [p.close_pos for p in profile.pairs]
    all_close_positions += [pos for _, pos in profile.unmatched_closes]
    if text[last] in CLOSE_QUOTES:
        if text[last] != target_close:
            return _replace_char(text, last, target_close), [f"右引号符号应由 {text[last]!r} 改为 {target_close!r}。"], True
        return text, [], True
    if all_close_positions:
        return text, ["译文存在右引号，但不在全文结尾，需手动调整位置。"], False
    return _insert_char(text, last + 1, target_close), [f"译文结尾缺少右引号 {target_close!r}。"], True


def _replace_existing_pair_symbols(
    text: str,
    target_open: str,
    target_close: str,
    require_end_close: bool | None,
) -> tuple[str, list[str], bool]:
    profile = _profile_translation_quotes(text)
    reasons: list[str] = []
    candidate: QuotePair | None = None
    if require_end_close is True:
        end = _last_visible(text)
        candidates = [p for p in profile.pairs if p.close_pos == end]
        if candidates:
            candidate = max(candidates, key=lambda p: p.close_pos - p.open_pos)
    elif profile.pairs:
        candidate = max(profile.pairs, key=lambda p: p.close_pos - p.open_pos)
    if candidate is None:
        return text, ["译文缺少可确认位置的配对引号，需手动处理。"], False
    result = text
    if candidate.close_char != target_close:
        reasons.append(f"右引号符号应由 {candidate.close_char!r} 改为 {target_close!r}。")
        result = _replace_char(result, candidate.close_pos, target_close)
    if candidate.open_char != target_open:
        reasons.append(f"左引号符号应由 {candidate.open_char!r} 改为 {target_open!r}。")
        result = _replace_char(result, candidate.open_pos, target_open)
    return result, reasons, True


def _fix_case2_quotes(text: str, target_open: str, target_close: str) -> tuple[str, list[str], bool]:
    """Case 2: opening quote is mid-text and closing quote must be at text end.

    The opening position may never be invented. An existing opening quote can be
    normalized in place. If that opening is unmatched, the missing closing quote
    can be added at the end. A closing quote that already exists away from the end
    is considered a position problem and therefore requires manual correction.
    """
    profile = _profile_translation_quotes(text)
    last = _last_visible(text)
    if last is None:
        return text, ["译文为空，无法处理引号。"], False

    open_candidates: list[tuple[str, int]] = []
    open_candidates.extend((p.open_char, p.open_pos) for p in profile.pairs)
    open_candidates.extend(profile.unmatched_opens)
    open_candidates = sorted(set(open_candidates), key=lambda item: item[1])
    if not open_candidates:
        return text, ["原文左引号位于文本中间，译文缺少左引号，需手动确定位置。"], False
    if len(open_candidates) > 1:
        return text, ["译文存在多个可能的左引号位置，无法安全判断对应位置，需手动处理。"], False

    open_char, open_pos = open_candidates[0]
    end_close = text[last] if text[last] in CLOSE_QUOTES else None
    non_end_closes = [p.close_pos for p in profile.pairs if p.close_pos != last]
    non_end_closes += [pos for _, pos in profile.unmatched_closes if pos != last]
    if end_close is None and non_end_closes:
        return text, ["译文已有右引号，但不在全文结尾，需手动调整位置。"], False

    result = text
    reasons: list[str] = []
    if end_close is not None and end_close != target_close:
        reasons.append(f"右引号符号应由 {end_close!r} 改为 {target_close!r}。")
        result = _replace_char(result, last, target_close)
    elif end_close is None:
        reasons.append(f"译文结尾缺少右引号 {target_close!r}。")
        result = _insert_char(result, last + 1, target_close)

    # Inserting at the end does not change the opening position.
    if open_char != target_open:
        reasons.append(f"左引号符号应由 {open_char!r} 改为 {target_open!r}。")
        result = _replace_char(result, open_pos, target_open)
    return result, reasons, True


def _apply_quote_rule(original: str, translated: str, profile: QuoteProfile) -> tuple[str, list[str], bool]:
    result = translated
    reasons: list[str] = []
    auto = True
    if profile.case_code == "none":
        return result, reasons, True
    if profile.case_code == "irregular":
        return result, ["原文引号结构不属于五种明确情形，需手动检查。"], False
    target_open = profile.open_char
    target_close = profile.close_char
    if profile.case_code == "case1":
        assert target_open and target_close and profile.open_line is not None
        result, rs, ok = _ensure_line_start_open(result, profile.open_line, target_open)
        reasons.extend(rs)
        auto &= ok
        if auto:
            result, rs, ok = _ensure_case1_end_close(result, target_close)
            reasons.extend(rs)
            auto &= ok
    elif profile.case_code == "case2":
        assert target_open and target_close
        result2, rs, ok = _fix_case2_quotes(result, target_open, target_close)
        if not ok:
            return translated, rs, False
        result = result2
        reasons.extend(rs)
    elif profile.case_code == "case3":
        assert target_open and target_close
        result2, rs, ok = _replace_existing_pair_symbols(result, target_open, target_close, False)
        if not ok:
            return translated, ["原文右引号不在全文结尾；译文缺少可对应的配对引号，需手动处理。"], False
        result = result2
        reasons.extend(rs)
    elif profile.case_code == "case4":
        assert target_open and target_close and profile.open_line is not None
        result, rs, ok = _ensure_line_start_open(result, profile.open_line, target_open)
        reasons.extend(rs)
        auto &= ok
        if auto:
            # 只删除与目标行首左引号相对应的右引号，避免误删内部嵌套引号。
            trans_profile = _profile_translation_quotes(result)
            target_open_pos = _first_visible_on_line(result, profile.open_line)
            target_pos = None
            if target_open_pos is not None:
                paired = [p.close_pos for p in trans_profile.pairs if p.open_pos == target_open_pos]
                if paired:
                    target_pos = max(paired)
            if target_pos is None:
                matching_unmatched = [
                    pos for char, pos in trans_profile.unmatched_closes if char == target_close
                ]
                if matching_unmatched:
                    target_pos = max(matching_unmatched)
            if target_pos is not None:
                reasons.append("译文存在原文不应有的右引号，需删除。")
                result = _delete_char(result, target_pos)
    elif profile.case_code == "case5":
        assert target_close
        trans_profile = _profile_translation_quotes(result)
        if trans_profile.pairs or trans_profile.unmatched_opens:
            return translated, ["原文只有结尾右引号，但译文含左引号，需手动判断。"], False
        result, rs, ok = _ensure_end_close(result, target_close)
        reasons.extend(rs)
        auto &= ok
    return result, reasons, auto


def build_dialogue_issue(record: DataRecord) -> DialogueIssue | None:
    if record.marker != "Message":
        return None
    original_profile = classify_quote_profile(record.original)
    quoted_result, quote_reasons, quote_auto = _apply_quote_rule(
        record.original, record.translated, original_profile
    )
    if quote_auto:
        final_result, punct_reasons = _apply_punctuation_rule(record.original, quoted_result)
    else:
        final_result = record.translated
        punct_reasons = []
        # 仍显示标点问题，但不能选择自动写回。
        _, punct_reasons = _apply_punctuation_rule(record.original, record.translated)
    reasons = quote_reasons + punct_reasons
    if not reasons:
        return None
    if original_profile.case_code == "none":
        category = "仅标点"
        case_code = "punctuation"
    elif original_profile.case_code == "irregular":
        category = "引号异常"
        case_code = "irregular"
    else:
        category = {
            "case1": "情形1",
            "case2": "情形2",
            "case3": "情形3",
            "case4": "情形4",
            "case5": "情形5",
        }[original_profile.case_code]
        case_code = original_profile.case_code
    proposed = final_result if quote_auto and final_result != record.translated else None
    return DialogueIssue(
        issue_id=f"dialogue|{record.uid}",
        record=record,
        case_code=case_code,
        category_label=category,
        reasons=tuple(reasons),
        current_translation=record.translated,
        proposed_translation=proposed,
        auto_fixable=quote_auto and proposed is not None,
        original_profile=original_profile,
    )


def analyze_dialogue_format(records: Iterable[DataRecord]) -> list[DialogueIssue]:
    return [issue for record in records if (issue := build_dialogue_issue(record)) is not None]


def _decode_csv(path: Path) -> str:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "cp932"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise QAError(f"无法识别 CSV 编码：{path}")


def _find_dictionary_columns(headers: Sequence[object]) -> tuple[int, int] | None:
    normalized = {_normalize_header(value): idx for idx, value in enumerate(headers) if value is not None}
    original_idx = next((normalized[h] for h in (_normalize_header(v) for v in ORIGINAL_HEADERS) if h in normalized), None)
    translation_idx = next((normalized[h] for h in (_normalize_header(v) for v in TRANSLATION_HEADERS) if h in normalized), None)
    if original_idx is None or translation_idx is None:
        return None
    return original_idx, translation_idx


def load_dictionary(path: Path) -> list[DictionaryEntry]:
    if not path.is_file():
        raise QAError(f"辞典文件不存在：{path}")
    rows: list[tuple[str, str]] = []
    if path.suffix.lower() == ".csv":
        reader = csv.reader(io.StringIO(_decode_csv(path)))
        all_rows = list(reader)
        if not all_rows:
            raise QAError(f"辞典为空：{path}")
        columns = _find_dictionary_columns(all_rows[0])
        if columns is None:
            raise QAError(f"辞典标题必须为 Original/Translation 或 原文/译文：{path}")
        o_col, t_col = columns
        for row in all_rows[1:]:
            original = row[o_col] if len(row) > o_col else ""
            translation = row[t_col] if len(row) > t_col else ""
            rows.append((str(original or ""), str(translation or "")))
    elif path.suffix.lower() in {".xlsx", ".xlsm"}:
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            found = False
            for ws in wb.worksheets:
                raw_max_row, max_col = _worksheet_dimensions(ws)
                max_row = min(raw_max_row, 200000)
                for header_row in range(1, min(max_row, 20) + 1):
                    headers = [ws.cell(header_row, col).value for col in range(1, max_col + 1)]
                    columns = _find_dictionary_columns(headers)
                    if columns is None:
                        continue
                    o_col, t_col = columns[0] + 1, columns[1] + 1
                    for row_num in range(header_row + 1, max_row + 1):
                        rows.append((str(ws.cell(row_num, o_col).value or ""), str(ws.cell(row_num, t_col).value or "")))
                    found = True
                    break
                if found:
                    break
            if not found:
                raise QAError(f"辞典标题必须为 Original/Translation 或 原文/译文：{path}")
        finally:
            wb.close()
    else:
        raise QAError("辞典仅支持 CSV、XLSX、XLSM。")

    mapping: OrderedDict[str, str] = OrderedDict()
    for index, (original, translation) in enumerate(rows, start=2):
        original = original.strip()
        translation = translation.strip()
        if not original and not translation:
            continue
        if not original or not translation:
            raise QAError(f"辞典 {path.name} 第 {index} 行原文或译文为空。")
        if original in mapping and mapping[original] != translation:
            raise QAError(
                f"辞典 {path.name} 中 {original!r} 对应多个译文：{mapping[original]!r} / {translation!r}"
            )
        mapping[original] = translation
    if not mapping:
        raise QAError(f"辞典中没有有效记录：{path}")
    return [DictionaryEntry(o, t, path) for o, t in mapping.items()]


def load_dictionaries(paths: Sequence[Path]) -> list[DictionaryEntry]:
    if not paths:
        raise QAError("尚未选择辞典。")
    combined: OrderedDict[str, DictionaryEntry] = OrderedDict()
    for path in paths:
        for entry in load_dictionary(path):
            previous = combined.get(entry.original)
            if previous and previous.translation != entry.translation:
                raise QAError(
                    f"多个辞典对 {entry.original!r} 给出不同译文："
                    f"{previous.translation!r}（{previous.source_file.name}）/ "
                    f"{entry.translation!r}（{entry.source_file.name}）"
                )
            combined.setdefault(entry.original, entry)
    return sorted(combined.values(), key=lambda e: (-len(e.original), e.original, str(e.source_file)))


def longest_non_overlapping_matches(text: str, entries: Sequence[DictionaryEntry]) -> list[DictionaryEntry]:
    candidates: list[tuple[int, int, DictionaryEntry]] = []
    for entry in entries:
        start = 0
        while True:
            pos = text.find(entry.original, start)
            if pos < 0:
                break
            candidates.append((pos, pos + len(entry.original), entry))
            start = pos + max(1, len(entry.original))
    candidates.sort(key=lambda item: (item[0], -(item[1] - item[0]), item[2].original))
    occupied: list[tuple[int, int]] = []
    selected: list[DictionaryEntry] = []
    for start, end, entry in candidates:
        if any(not (end <= a or start >= b) for a, b in occupied):
            continue
        occupied.append((start, end))
        selected.append(entry)
    return selected


def analyze_dictionary_matches(
    records: Iterable[DataRecord],
    entries: Sequence[DictionaryEntry],
    *,
    messages_only: bool = False,
) -> list[DictionaryWarning]:
    results: list[DictionaryWarning] = []
    for record in records:
        if messages_only and record.marker != "Message":
            continue
        original_visible = strip_control_codes(record.original)
        translated_visible = strip_control_codes(record.translated)
        seen: set[tuple[str, str]] = set()
        for entry in longest_non_overlapping_matches(original_visible, entries):
            pair = (entry.original, entry.translation)
            if pair in seen:
                continue
            seen.add(pair)
            if entry.translation not in translated_visible:
                results.append(
                    DictionaryWarning(
                        record=record,
                        original_term=entry.original,
                        expected_translation=entry.translation,
                        source_dictionary=entry.source_file,
                        original_visible=original_visible,
                        translated_visible=translated_visible,
                    )
                )
    return results


def normalize_for_search(text: str) -> str:
    text = unicodedata.normalize("NFKC", strip_control_codes(text)).casefold()
    return "".join(
        char for char in text
        if not unicodedata.category(char).startswith(("P", "S", "Z", "C"))
    )


def search_records(
    records: Iterable[DataRecord],
    query: str,
    *,
    scope: str = "both",  # original/translated/both
    mode: str = "keyword",  # keyword/full
) -> list[SearchHit]:
    normalized_query = normalize_for_search(query)
    if not normalized_query:
        raise QAError("查询内容去除控制符和特殊符号后为空。")
    results: list[SearchHit] = []
    for record in records:
        original_norm = normalize_for_search(record.original)
        translated_norm = normalize_for_search(record.translated)
        match_original = scope in {"original", "both"} and (
            normalized_query in original_norm if mode == "keyword" else normalized_query == original_norm
        )
        match_translated = scope in {"translated", "both"} and (
            normalized_query in translated_norm if mode == "keyword" else normalized_query == translated_norm
        )
        if not match_original and not match_translated:
            continue
        matched_field = "原文和译文" if match_original and match_translated else ("原文" if match_original else "译文")
        results.append(
            SearchHit(
                record=record,
                matched_field=matched_field,
                original_excerpt=_compact_excerpt(record.original, query),
                translated_excerpt=_compact_excerpt(record.translated, query),
            )
        )
    return results


def _compact_excerpt(text: str, query: str, limit: int = 180) -> str:
    visible = strip_control_codes(normalize_newlines(text)).replace("\n", " ↵ ")
    if len(visible) <= limit:
        return visible
    normalized_query = normalize_for_search(query)
    if not normalized_query:
        return visible[: limit - 1] + "…"
    # 特殊符号被移除后无法精确回映射，优先在原显示文本中做普通查找。
    pos = visible.casefold().find(query.casefold())
    if pos < 0:
        return visible[: limit - 1] + "…"
    start = max(0, pos - limit // 3)
    end = min(len(visible), start + limit)
    prefix = "…" if start else ""
    suffix = "…" if end < len(visible) else ""
    return prefix + visible[start:end] + suffix


def generate_specialized_dictionary(
    records: Iterable[DataRecord],
    entries: Sequence[DictionaryEntry],
) -> list[GeneratedDictionaryRow]:
    counts: Counter[tuple[str, str, str]] = Counter()
    first_order: list[tuple[str, str, str]] = []
    for record in records:
        visible = strip_control_codes(record.original)
        for entry in longest_non_overlapping_matches(visible, entries):
            key = (entry.original, entry.translation, entry.source_file.name)
            if key not in counts:
                first_order.append(key)
            counts[key] += 1
    keys = sorted(counts, key=lambda item: (item[2].casefold(), -len(item[0]), item[0], item[1]))
    return [GeneratedDictionaryRow(o, t, source, counts[(o, t, source)]) for o, t, source in keys]


def save_specialized_dictionary(path: Path, rows: Sequence[GeneratedDictionaryRow]) -> None:
    if not rows:
        raise QAError("原文没有匹配到任何辞典词条。")
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.suffix.lower() == ".csv":
        with path.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.writer(fh)
            writer.writerow(["Original", "Translation", "SourceDictionary", "MatchCount"])
            for row in rows:
                writer.writerow([row.original, row.translation, row.source_dictionary, row.match_count])
        return
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise QAError("输出辞典请选择 .csv 或 .xlsx。")
    wb = Workbook()
    fixed_dt = datetime(2000, 1, 1, 0, 0, 0)
    wb.properties.creator = APP_NAME
    wb.properties.lastModifiedBy = APP_NAME
    wb.properties.created = fixed_dt
    wb.properties.modified = fixed_dt
    ws = wb.active
    ws.title = "Dictionary"
    headers = ["Original", "Translation", "SourceDictionary", "MatchCount"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
    for row_num, row in enumerate(rows, start=2):
        ws.cell(row_num, 1, row.original)
        ws.cell(row_num, 2, row.translation)
        ws.cell(row_num, 3, row.source_dictionary)
        ws.cell(row_num, 4, row.match_count)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{len(rows) + 1}"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 12
    temp = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
    wb.save(temp)
    wb.close()
    _normalize_xlsx_package(temp)
    _validate_xlsx(temp)
    os.replace(temp, path)


def _timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _normalize_xlsx_package(path: Path) -> None:
    """Normalize ZIP order and timestamps so equal dictionary data yields equal XLSX bytes."""
    fixed_time = (1980, 1, 1, 0, 0, 0)
    normalized = path.with_name(f".{path.name}.{uuid.uuid4().hex}.normalized")
    try:
        with zipfile.ZipFile(path, "r") as src:
            entries = [
                (info.filename, src.read(info.filename), info.compress_type, info.external_attr)
                for info in src.infolist()
            ]
        with zipfile.ZipFile(normalized, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9) as dst:
            for name, payload, compress_type, external_attr in sorted(entries, key=lambda item: item[0]):
                info = zipfile.ZipInfo(name, date_time=fixed_time)
                info.compress_type = compress_type
                info.create_system = 0
                info.external_attr = external_attr
                dst.writestr(info, payload)
        os.replace(normalized, path)
    finally:
        if normalized.exists():
            normalized.unlink(missing_ok=True)


def _validate_xlsx(path: Path) -> None:
    with zipfile.ZipFile(path, "r") as archive:
        bad = archive.testzip()
        if bad:
            raise QAError(f"Excel 文件 ZIP 结构损坏：{bad}")
        names = set(archive.namelist())
        if "[Content_Types].xml" not in names or "xl/workbook.xml" not in names:
            raise QAError("Excel 文件缺少必要 XML 部件。")


def _atomic_write_json(path: Path, data: Mapping) -> None:
    temp = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
    with temp.open("w", encoding="utf-8", newline="\n") as fh:
        json.dump(data, fh, ensure_ascii=False, indent=4)
        fh.flush()
        os.fsync(fh.fileno())
    os.replace(temp, path)


def _backup_root(base: Path) -> Path:
    root = base / BACKUP_DIRNAME / f"QA_{_timestamp()}"
    suffix = 1
    while root.exists():
        root = base / BACKUP_DIRNAME / f"QA_{_timestamp()}_{suffix}"
        suffix += 1
    return root


def apply_dialogue_fixes(
    selected: Sequence[DialogueIssue],
    *,
    json_path: Path | None = None,
    translated_dir: Path | None = None,
    table_dir: Path | None = None,
    log: LogFn = _noop_log,
) -> tuple[int, Path]:
    issues = [issue for issue in selected if issue.auto_fixable and issue.proposed_translation is not None]
    if not issues:
        raise QAError("没有选中可自动修复的项目。")
    kinds = {issue.record.source_kind for issue in issues}
    if len(kinds) != 1:
        raise QAError("一次修复只能处理同一种数据源。")
    kind = next(iter(kinds))
    if kind == "json":
        if json_path is None:
            json_path = issues[0].record.translated_open_path
        backup_root = _backup_root(json_path.parent)
        backup_root.mkdir(parents=True, exist_ok=True)
        shutil.copy2(json_path, backup_root / json_path.name)
        with json_path.open("r", encoding="utf-8-sig") as fh:
            data = json.load(fh, object_pairs_hook=_no_duplicate_object)
        for issue in issues:
            file_key, original_key = issue.record.update_ref
            current = data[file_key][original_key]["text"]
            if current != issue.current_translation:
                raise QAError(f"JSON 在扫描后已发生变化，停止覆盖：{file_key} / {original_key[:40]!r}")
            data[file_key][original_key]["text"] = issue.proposed_translation
        _atomic_write_json(json_path, data)
        log(f"已修复 JSON：{len(issues)} 条")
        return len(issues), backup_root
    if kind == "excel":
        if table_dir is None:
            table_dir = Path(str(issues[0].record.metadata.get("table_dir") or issues[0].record.translated_open_path.parent))
        backup_root = _backup_root(table_dir)
        grouped: dict[Path, list[DialogueIssue]] = {}
        for issue in issues:
            grouped.setdefault(issue.record.translated_open_path, []).append(issue)
        applied = 0
        for path, file_issues in grouped.items():
            try:
                relative = path.relative_to(table_dir)
            except ValueError:
                relative = Path(path.name)
            backup_path = backup_root / relative
            backup_path.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(path, backup_path)
            keep_vba = path.suffix.lower() == ".xlsm"
            wb = load_workbook(path, read_only=False, data_only=False, keep_vba=keep_vba)
            try:
                for issue in file_issues:
                    _, sheet, row, col = issue.record.update_ref
                    ws = wb[sheet]
                    cell = ws.cell(row, col)
                    current = "" if cell.value is None else str(cell.value)
                    if current != issue.current_translation:
                        raise QAError(f"Excel 在扫描后已发生变化，停止覆盖：{path} / {sheet}!{cell.coordinate}")
                    cell.value = issue.proposed_translation
                    cell.data_type = "s"
                    applied += 1
                temp = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp{path.suffix}")
                wb.save(temp)
            finally:
                wb.close()
            _validate_xlsx(temp)
            os.replace(temp, path)
            log(f"已修复 Excel：{path.name}（{len(file_issues)} 条）")
        return applied, backup_root
    if kind == "txt":
        if translated_dir is None:
            # 从第一个目标文件向上推断不可靠，GUI 必须明确传入。
            raise QAError("修复 TXT 时必须提供译文 TXT 文件夹。")
        backup_root = _backup_root(translated_dir.parent)
        grouped: dict[Path, list[DialogueIssue]] = {}
        for issue in issues:
            grouped.setdefault(issue.record.translated_open_path, []).append(issue)
        applied = 0
        for path, file_issues in grouped.items():
            if not path.exists():
                original_path = file_issues[0].record.original_open_path
                path.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(original_path, path)
            try:
                relative = path.relative_to(translated_dir)
            except ValueError:
                relative = Path(path.name)
            backup_path = backup_root / translated_dir.name / relative
            backup_path.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(path, backup_path)
            current_occ = {occ.locator: occ for occ in parse_string_script(path)}
            text, has_bom, newline = _read_text_file(path)
            lines = text.splitlines(keepends=True)
            replacements: list[tuple[int, int, list[str]]] = []
            for issue in file_issues:
                _, locator, kind_hint = issue.record.update_ref
                occ = current_occ.get(locator)
                if occ is None:
                    raise QAError(f"译文 TXT 中找不到扫描时条目：{path} / {locator}")
                if occ.text != issue.current_translation:
                    raise QAError(f"TXT 在扫描后已发生变化，停止覆盖：{path} / {locator}")
                replacement = _format_occurrence_replacement(lines, occ, issue.proposed_translation or "")
                replacements.append((occ.start_line, occ.end_line, replacement))
                applied += 1
            for start, end, replacement in sorted(replacements, key=lambda item: item[0], reverse=True):
                lines[start:end] = replacement
            temp = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
            _write_text_file(temp, "".join(lines), has_bom=has_bom, newline=newline)
            os.replace(temp, path)
            log(f"已修复 TXT：{relative}（{len(file_issues)} 条）")
        return applied, backup_root
    raise QAError(f"不支持的数据源：{kind}")


def _format_occurrence_replacement(lines: list[str], occ: ParsedOccurrence, translated: str) -> list[str]:
    translated = normalize_newlines(translated)
    if occ.kind == "multiline":
        original_block = "".join(lines[occ.start_line:occ.end_line])
        replacement = translated
        if original_block.endswith("\n") and not replacement.endswith("\n"):
            replacement += "\n"
        elif not original_block.endswith("\n") and replacement.endswith("\n"):
            replacement = replacement.rstrip("\n")
        result = replacement.splitlines(keepends=True)
        return result if result else ([replacement] if replacement else [])
    raw = lines[occ.start_line].rstrip("\n")
    leading_len = len(raw) - len(raw.lstrip())
    trailing_len = len(raw) - len(raw.rstrip())
    leading = raw[:leading_len]
    trailing = raw[len(raw) - trailing_len:] if trailing_len else ""
    return [f"{leading}{translated.strip()}{trailing}\n"]


def open_file(path: Path) -> str:
    path = path.resolve()
    if not path.exists():
        raise QAError(f"文件不存在：{path}")
    try:
        if sys.platform == "win32":
            os.startfile(str(path))  # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            subprocess.Popen(["open", str(path)])
        else:
            subprocess.Popen(["xdg-open", str(path)])
    except Exception as exc:
        raise QAError(f"无法打开文件：{exc}") from exc
    return f"已打开：{path}"


def save_config(path: Path, data: Mapping) -> None:
    temp = path.with_suffix(path.suffix + ".tmp")
    temp.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(temp, path)


def load_config(path: Path) -> dict:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}
